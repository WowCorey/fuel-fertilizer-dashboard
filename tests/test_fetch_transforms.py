import json
import pathlib
import sys
import tempfile
import unittest
import datetime as dt
from unittest import mock

ROOT = pathlib.Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from scripts import fetch_data


class FakeResponse:
    def __init__(self, *, text="", content=None, json_doc=None, status_code=200):
        self.text = text
        self.content = content if content is not None else text.encode("utf-8")
        self._json_doc = json_doc
        self.status_code = status_code

    def raise_for_status(self):
        if self.status_code >= 400:
            raise RuntimeError(f"HTTP {self.status_code}")

    def json(self):
        return self._json_doc


class FetchTransformTests(unittest.TestCase):
    def test_non_required_fetch_check_failure_does_not_block(self):
        source = {
            "id": "slow_optional_source",
            "fetch": "programmatic",
            "fetch_url": "https://example.test/slow.xlsx",
            "check_required": False,
        }
        with mock.patch.object(fetch_data, "load_sources", return_value=[source]), \
             mock.patch.object(fetch_data, "check_url", return_value=(False, "ReadTimeout: timed out")), \
             mock.patch.object(sys, "argv", ["fetch_data.py", "--check"]):
            self.assertEqual(fetch_data.main(), 0)

    def test_required_fetch_check_failure_blocks(self):
        source = {
            "id": "required_source",
            "fetch": "programmatic",
            "fetch_url": "https://example.test/required.csv",
        }
        with mock.patch.object(fetch_data, "load_sources", return_value=[source]), \
             mock.patch.object(fetch_data, "check_url", return_value=(False, "ReadTimeout: timed out")), \
             mock.patch.object(sys, "argv", ["fetch_data.py", "--check"]):
            self.assertEqual(fetch_data.main(), 1)

    def test_non_required_fetch_failure_keeps_existing_generated_envelope(self):
        source = {
            "id": "slow_optional_source",
            "human_name": "Slow optional source",
            "fetch": "programmatic",
            "fetch_url": "https://example.test/slow.xlsx",
            "fetch_required": False,
        }

        def fail_fetch(_source):
            raise fetch_data.requests.exceptions.ReadTimeout("timed out")

        with tempfile.TemporaryDirectory() as tmp:
            generated = pathlib.Path(tmp)
            (generated / "slow_optional_source.json").write_text("{}", encoding="utf-8")
            old_generated = fetch_data.GENERATED_DIR
            fetch_data.GENERATED_DIR = generated
            try:
                with mock.patch.object(fetch_data, "load_sources", return_value=[source]), \
                     mock.patch.dict(fetch_data.FETCHERS, {"slow_optional_source": fail_fetch}), \
                     mock.patch.object(sys, "argv", ["fetch_data.py"]):
                    self.assertEqual(fetch_data.main(), 0)
            finally:
                fetch_data.GENERATED_DIR = old_generated

    def test_non_required_fetch_failure_without_existing_envelope_blocks(self):
        source = {
            "id": "slow_optional_source",
            "human_name": "Slow optional source",
            "fetch": "programmatic",
            "fetch_url": "https://example.test/slow.xlsx",
            "fetch_required": False,
        }

        def fail_fetch(_source):
            raise fetch_data.requests.exceptions.ReadTimeout("timed out")

        with tempfile.TemporaryDirectory() as tmp:
            old_generated = fetch_data.GENERATED_DIR
            fetch_data.GENERATED_DIR = pathlib.Path(tmp)
            try:
                with mock.patch.object(fetch_data, "load_sources", return_value=[source]), \
                     mock.patch.dict(fetch_data.FETCHERS, {"slow_optional_source": fail_fetch}), \
                     mock.patch.object(sys, "argv", ["fetch_data.py"]):
                    self.assertEqual(fetch_data.main(), 1)
            finally:
                fetch_data.GENERATED_DIR = old_generated

    def test_fetch_failure_detail_includes_source_context(self):
        source = {
            "id": "slow_optional_source",
            "human_name": "Slow optional source",
            "fetch_url": "https://example.test/slow.xlsx",
        }
        detail = fetch_data.fetch_failure_detail(
            source,
            fetch_data.requests.exceptions.ReadTimeout("timed out"),
            required=False,
        )
        self.assertIn("slow_optional_source", detail)
        self.assertIn("Slow optional source", detail)
        self.assertIn("https://example.test/slow.xlsx", detail)
        self.assertIn("fetch_required=False", detail)
        self.assertIn("non-blocking", detail)

    def test_abs_sdmx_parses_single_series(self):
        doc = {
            "dataSets": [
                {
                    "series": {
                        "0:0:0": {
                            "observations": {
                                "0": [1000],
                                "1": [1250],
                            }
                        }
                    }
                }
            ],
            "structure": {
                "dimensions": {
                    "observation": [
                        {
                            "id": "TIME_PERIOD",
                            "values": [
                                {"id": "2026-01", "end": "2026-01-31T00:00:00"},
                                {"id": "2026-02", "end": "2026-02-28T00:00:00"},
                            ],
                        }
                    ]
                },
                "attributes": {
                    "dataSet": [
                        {"id": "UNIT_MEASURE", "values": [{"id": "AUD"}]},
                        {"id": "UNIT_MULT", "values": [{"id": "3"}]},
                    ]
                },
            },
        }
        with mock.patch.object(fetch_data.requests, "get", return_value=FakeResponse(json_doc=doc)) as get:
            out = fetch_data.fetch_abs_sdmx("MERCH_IMP", "33.TOT.TOT.M", "2026-01")

        self.assertEqual(
            get.call_args.args[0],
            "https://data.api.abs.gov.au/rest/data/MERCH_IMP/33.TOT.TOT.M?format=jsondata&startPeriod=2026-01",
        )
        self.assertEqual(out["unit"], "AUD thousands")
        self.assertEqual(out["last_data_point"], "2026-02-28")
        self.assertEqual(out["values"], [{"t": "2026-01", "v": 1000}, {"t": "2026-02", "v": 1250}])

    def test_petroleum_imports_yoy_derivation(self):
        source = {"id": "abs_petroleum_imports_yoy", "derived_from": "abs_petroleum_imports"}
        with tempfile.TemporaryDirectory() as tmp:
            generated = pathlib.Path(tmp)
            parent = {
                "status": "ok",
                "last_data_point": "2026-01-31",
                "source_url": "https://example.test/source",
                "values": [
                    {"t": f"2025-{month:02d}", "v": 100.0}
                    for month in range(1, 13)
                ] + [{"t": "2026-01", "v": 125.0}],
            }
            (generated / "abs_petroleum_imports.json").write_text(json.dumps(parent), encoding="utf-8")
            old_generated = fetch_data.GENERATED_DIR
            fetch_data.GENERATED_DIR = generated
            try:
                out = fetch_data.fetch_abs_petroleum_imports_yoy(source)
            finally:
                fetch_data.GENERATED_DIR = old_generated

        self.assertEqual(out["unit"], "%")
        self.assertEqual(out["values"][-1], {"t": "2026-01", "v": 25.0})
        self.assertEqual(out["last_data_point"], "2026-01-31")

    def test_extra_field_derivation_reads_manual_parent(self):
        source = {
            "id": "fuel_security_petrol_days_remaining",
            "derived_from": "pmc_mso_days_cover",
            "derived_field": "petrol_days",
            "derived_label": "Petrol",
            "fetch_unit": "days",
        }
        with tempfile.TemporaryDirectory() as tmp:
            manual = pathlib.Path(tmp)
            parent = {
                "status": "ok",
                "manual_entry": True,
                "last_data_point": "2026-04-14",
                "source_url": "https://example.test/pmc",
                "unit": "days",
                "values": [{"t": "2026-04-14", "v": 30}],
                "extra": {
                    "schema": "pmc_mso_days_cover.v1",
                    "fields": {
                        "as_at": "2026-04-14",
                        "petrol_days": 46,
                    },
                },
            }
            (manual / "pmc_mso_days_cover.json").write_text(json.dumps(parent), encoding="utf-8")
            old_manual = fetch_data.MANUAL_DIR
            old_generated = fetch_data.GENERATED_DIR
            fetch_data.MANUAL_DIR = manual
            fetch_data.GENERATED_DIR = pathlib.Path(tmp) / "generated"
            try:
                out = fetch_data.fetch_extra_field_derived(source)
            finally:
                fetch_data.MANUAL_DIR = old_manual
                fetch_data.GENERATED_DIR = old_generated

        self.assertEqual(out["unit"], "days")
        self.assertEqual(out["values"], [{"t": "2026-04-14", "v": 46.0}])
        self.assertEqual(out["last_data_point"], "2026-04-14")
        self.assertEqual(out["extra"]["fields"]["parent_source_id"], "pmc_mso_days_cover")
        self.assertTrue(out["extra"]["fields"]["parent_manual_entry"])

    def test_abs_fertiliser_source_concentration(self):
        countries = [
            ("TOT", "Total"),
            ("SA", "Saudi Arabia"),
            ("MA", "Morocco"),
            ("CN", "China"),
            ("US", "United States"),
        ]
        doc = {
            "dataSets": [{"series": {}}],
            "structure": {
                "dimensions": {
                    "series": [
                        {"id": "COMMODITY_SITC", "values": [{"id": "562", "name": "Fertilisers"}]},
                        {"id": "COUNTRY_ORIGIN", "values": [{"id": code, "name": name} for code, name in countries]},
                        {"id": "STATE_DEST", "values": [{"id": "TOT", "name": "Total"}]},
                        {"id": "FREQ", "values": [{"id": "M", "name": "Monthly"}]},
                    ],
                    "observation": [
                        {
                            "id": "TIME_PERIOD",
                            "values": [
                                {"id": "2026-01", "end": "2026-01-31T00:00:00"},
                                {"id": "2026-02", "end": "2026-02-28T00:00:00"},
                            ],
                        }
                    ],
                }
            },
        }
        rows = {
            0: [1000, 2000],
            1: [300, 800],
            2: [250, 500],
            3: [200, 400],
            4: [100, 100],
        }
        for country_idx, values in rows.items():
            doc["dataSets"][0]["series"][f"0:{country_idx}:0:0"] = {
                "observations": {"0": [values[0]], "1": [values[1]]}
            }

        source = {
            "id": "abs_fertiliser_source_concentration",
            "fetch_dataflow": "MERCH_IMP",
            "fetch_key": "562..TOT.M",
            "fetch_start_period": "2026-01",
            "fetch_commodity": "562",
        }
        with mock.patch.object(fetch_data.requests, "get", return_value=FakeResponse(json_doc=doc)) as get:
            out = fetch_data.fetch_abs_fertiliser_source_concentration(source)

        self.assertEqual(
            get.call_args.args[0],
            "https://data.api.abs.gov.au/rest/data/MERCH_IMP/562..TOT.M?format=jsondata&startPeriod=2026-01",
        )
        self.assertEqual(out["unit"], "%")
        self.assertEqual(out["last_data_point"], "2026-02-28")
        self.assertEqual(out["values"], [{"t": "2026-01", "v": 75.0}, {"t": "2026-02", "v": 85.0}])
        fields = out["extra"]["fields"]
        self.assertEqual(fields["top3_share_percent"], 85.0)
        self.assertEqual(fields["total_import_value_aud_thousands"], 2000)
        self.assertEqual([row["country"] for row in fields["top_countries"]], ["Saudi Arabia", "Morocco", "China"])

    def test_rba_f11_monthly_means(self):
        csv_text = "\n".join(
            [
                "Some intro",
                "More intro",
                "More intro",
                "More intro",
                "More intro",
                "More intro",
                "More intro",
                "More intro",
                "More intro",
                "Title,A$1=USD,Other",
                "Description,Australian dollar US dollar,Other",
                "Series ID,F11.1,Other",
                "01-Jan-2026,0.6500,x",
                "02-Jan-2026,0.6700,x",
                "03-Feb-2026,0.7000,x",
            ]
        )
        with mock.patch.object(fetch_data.requests, "get", return_value=FakeResponse(content=csv_text.encode("utf-8-sig"))):
            out = fetch_data.fetch_rba_f11("https://example.test/f11.csv")

        self.assertEqual(out["unit"], "USD per AUD")
        self.assertEqual(out["values"], [{"t": "2026-01", "v": 0.66}, {"t": "2026-02", "v": 0.7}])
        self.assertEqual(out["last_data_point"], "2026-02-03")

    def test_rba_csv_column_accepts_slash_dates(self):
        csv_text = "\n".join(
            [
                "Some intro",
                "More intro",
                "More intro",
                "More intro",
                "More intro",
                "More intro",
                "More intro",
                "More intro",
                "More intro",
                "Title,Cash Rate Target,Other",
                "Description,Cash rate target,Other",
                "Series ID,FIRMMCRT,Other",
                "01/01/2026,4.00,x",
                "31/01/2026,3.50,x",
            ]
        )
        with mock.patch.object(fetch_data.requests, "get", return_value=FakeResponse(content=csv_text.encode("utf-8-sig"))):
            out = fetch_data.fetch_rba_csv_column(
                "https://example.test/f1.1.csv",
                title_substring="Cash Rate Target",
                aggregate="monthly_mean",
                unit="per cent",
                notes="test",
                ndigits=2,
            )

        self.assertEqual(out["unit"], "per cent")
        self.assertEqual(out["values"], [{"t": "2026-01", "v": 3.75}])
        self.assertEqual(out["last_data_point"], "2026-01-31")

    def test_rba_cash_rate_latest_preserves_latest_raw_observation(self):
        html_text = """
            <table>
              <tr><th>Effective Date</th><th>Change %&nbsp;points</th><th>Cash rate target %</th><th>Related Documents</th></tr>
              <tr><td>6 May 2026</td><td>+0.25</td><td>4.35</td><td>Statement</td></tr>
              <tr><td>18 Mar 2026</td><td>+0.25</td><td>4.10</td><td>Statement</td></tr>
            </table>
        """
        source = {"id": "rba_cash_rate_latest", "fetch_url": "https://example.test/cash-rate/"}
        with mock.patch.object(fetch_data.requests, "get", return_value=FakeResponse(text=html_text)):
            out = fetch_data.fetch_rba_cash_rate_latest(source)

        self.assertEqual(out["unit"], "per cent")
        self.assertEqual(out["values"], [{"t": "2026-05-06", "v": 4.35}])
        self.assertEqual(out["last_data_point"], "2026-05-06")
        fields = out["extra"]["fields"]
        self.assertEqual(fields["latest_target_value"], 4.35)
        self.assertEqual(fields["previous_target_value"], 4.1)
        self.assertEqual(fields["previous_observation_date"], "2026-03-18")
        self.assertEqual(fields["target_change_percentage_points"], 0.25)
        self.assertEqual(fields["change_direction"], "up")

    def test_aofm_ags_face_value_csv(self):
        csv_text = "\n".join(
            [
                "FY,Face Value ($b)",
                "2023-24,878.9",
                "2024-25,895.6",
            ]
        )
        source = {"id": "aofm_gov_gross_debt", "fetch_url": "https://example.test/stock_ags.csv"}
        with mock.patch.object(fetch_data.requests, "get", return_value=FakeResponse(content=csv_text.encode("utf-8-sig"))):
            out = fetch_data.fetch_aofm_ags_face_value(source)

        self.assertEqual(out["unit"], "AUD billions")
        self.assertEqual(out["values"], [{"t": "2023-24", "v": 878.9}, {"t": "2024-25", "v": 895.6}])
        self.assertEqual(out["last_data_point"], "2025-06-30")
        self.assertIn("not the same as Commonwealth general government gross debt", out["notes"])

    def test_aps_xlsx_series_reads_named_sheet_and_column(self):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales of products"
        ws.append(["Month", "Diesel oil: total (ML)"])
        ws.append([dt.date(2026, 1, 1), 1200.44])
        ws.append([dt.date(2026, 2, 1), 1300.56])

        source = {
            "id": "aps_sales_diesel",
            "fetch_url": "https://example.test/package",
            "aps_sheet": "Sales of products",
            "aps_column": "Diesel oil: total (ML)",
            "fetch_unit": "ML",
        }
        with mock.patch.object(fetch_data, "discover_aps_workbook_url", return_value="https://example.test/aps.xlsx"), \
             mock.patch.object(fetch_data, "load_xlsx_from_url", return_value=wb):
            out = fetch_data.fetch_aps_xlsx_series(source)

        self.assertEqual(out["unit"], "ML")
        self.assertEqual(out["values"], [{"t": "2026-01", "v": 1200.4}, {"t": "2026-02", "v": 1300.6}])
        self.assertEqual(out["last_data_point"], "2026-02-28")
        self.assertIn("Sales of products", out["notes"])

    def test_aip_tgp_monthly_average_from_workbook(self):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Petrol TGP"
        ws.append(["Date", "National Average"])
        ws.append([dt.date(2026, 1, 1), 180.0])
        ws.append([dt.date(2026, 1, 2), 184.0])
        ws.append([dt.date(2026, 2, 1), 190.0])

        source = {"id": "aip_tgp", "fetch_url": "https://example.test/aip"}
        with mock.patch.object(fetch_data, "discover_aip_tgp_workbook_url", return_value="https://example.test/aip.xlsx"), \
             mock.patch.object(fetch_data, "load_xlsx_from_url", return_value=wb):
            out = fetch_data.fetch_aip_tgp(source)

        self.assertEqual(out["unit"], "cents per litre")
        self.assertEqual(out["values"], [{"t": "2026-01", "v": 182.0}, {"t": "2026-02", "v": 190.0}])
        self.assertEqual(out["last_data_point"], "2026-02-01")

    def test_qld_unavailable_reports_counts_price_9999(self):
        def fake_get(url, **_kwargs):
            if "package_show" in url:
                return FakeResponse(
                    json_doc={
                        "success": True,
                        "result": {
                            "resources": [
                                {
                                    "id": "res-jan",
                                    "name": "Queensland Fuel Prices January 2026",
                                    "format": "CSV",
                                    "datastore_active": True,
                                    "url": "https://example.test/jan.csv",
                                },
                                {
                                    "id": "res-feb",
                                    "name": "Queensland Fuel Prices February 2026",
                                    "format": "CSV",
                                    "datastore_active": True,
                                    "url": "https://example.test/feb.csv",
                                },
                            ]
                        },
                    }
                )
            sql = url
            if "GROUP+BY" in sql or "GROUP%20BY" in sql:
                return FakeResponse(
                    json_doc={
                        "success": True,
                        "result": {
                            "records": [
                                {"Fuel_Type": "Unleaded", "reports": 2, "sites": 2},
                                {"Fuel_Type": "Diesel", "reports": 1, "sites": 1},
                            ]
                        },
                    }
                )
            if "res-jan" in sql:
                record = {
                    "reports": 1,
                    "sites": 1,
                    "latest": "2026-01-12T00:00:00",
                    "earliest": "2026-01-12T00:00:00",
                }
            else:
                record = {
                    "reports": 3,
                    "sites": 2,
                    "latest": "2026-02-10T01:00:00",
                    "earliest": "2026-02-01T00:00:00",
                }
            return FakeResponse(json_doc={"success": True, "result": {"records": [record]}})

        source = {
            "id": "qld_fuel_security_unavailable_reports",
            "qld_ckan_base": "https://example.test",
            "qld_package_id": "fuel-price-reporting-2026",
        }
        with mock.patch.object(fetch_data.requests, "get", side_effect=fake_get):
            out = fetch_data.fetch_qld_unavailable_reports(source)

        self.assertEqual(out["unit"], "unavailable fuel-type reports")
        self.assertEqual(out["values"], [{"t": "2026-01", "v": 1}, {"t": "2026-02", "v": 3}])
        self.assertEqual(out["last_data_point"], "2026-02-10")
        fields = out["extra"]["fields"]
        self.assertEqual(fields["latest_resource_id"], "res-feb")
        self.assertEqual(fields["distinct_sites_with_unavailable_fuel"], 2)
        self.assertEqual(fields["fuel_type_breakdown"][0]["fuel_type"], "Unleaded")

    def test_retail_multistate_weighted_average(self):
        contributors = [
            {"state": "NSW", "average": 190.0, "stations": 10, "date": "2026-04-20"},
            {"state": "QLD", "average": 200.0, "stations": 20, "date": "2026-04-20"},
            {"state": "WA", "average": 220.0, "stations": 5, "date": "2026-04-20"},
        ]
        with mock.patch.object(fetch_data, "fetch_nsw_fuelcheck", return_value=contributors[0]), \
             mock.patch.object(fetch_data, "fetch_qld_open_data", return_value=contributors[1]), \
             mock.patch.object(fetch_data, "fetch_wa_fuelwatch", return_value=contributors[2]):
            out = fetch_data.fetch_retail_multistate({})

        self.assertEqual(out["unit"], "cents per litre")
        self.assertRegex(out["last_data_point"], r"^20\d{2}-\d{2}-\d{2}$")
        self.assertEqual(out["values"][0]["v"], 200.0)
        self.assertEqual(out["extra"]["fields"]["states"], contributors)

    def test_nopta_petroleum_counts_keep_object_classes_separate(self):
        calls = []

        def fake_get(url, params=None, **_kwargs):
            calls.append((url, params))
            if "TitlesCompany" in url:
                return FakeResponse(
                    json_doc={
                        "features": [
                            {
                                "attributes": {
                                    "OffShoreAr": "Western Australia",
                                    "TitleType": "Production Licence",
                                    "Status": "Active",
                                    "record_count": 2,
                                }
                            },
                            {
                                "attributes": {
                                    "OffShoreAr": "Western Australia",
                                    "TitleType": "Exploration Permit",
                                    "Status": "Pending Application",
                                    "record_count": 1,
                                }
                            },
                            {
                                "attributes": {
                                    "OffShoreAr": "Territory of Ashmore and Cartier Islands",
                                    "TitleType": "Retention Lease",
                                    "Status": "Active",
                                    "record_count": 3,
                                }
                            },
                        ]
                    }
                )
            return FakeResponse(
                json_doc={
                    "features": [
                        {
                            "attributes": {
                                "OffshoreArea": "Western Australia",
                                "Type": "Petroleum",
                                "IsOffshore": "Yes",
                                "record_count": 5,
                            }
                        },
                        {
                            "attributes": {
                                "OffshoreArea": "Queensland",
                                "Type": None,
                                "IsOffshore": "No",
                                "record_count": 7,
                            }
                        },
                        {
                            "attributes": {
                                "OffshoreArea": "Outside of Australia",
                                "Type": "Petroleum",
                                "IsOffshore": "Yes",
                                "record_count": 1,
                            }
                        },
                    ]
                }
            )

        source = {
            "id": "state_petroleum_nopta_counts",
            "fetch_url": "https://example.test/TitlesCompany/query",
            "nopta_wells_query_url": "https://example.test/Petroleum_Wells/query",
        }
        with mock.patch.object(fetch_data.requests, "get", side_effect=fake_get):
            out = fetch_data.fetch_nopta_petroleum_counts(source)

        self.assertEqual(out["unit"], "structured counts")
        self.assertEqual(out["values"], [])
        fields = out["extra"]["fields"]
        wa = next(row for row in fields["state_rows"] if row["state_code"] == "WA")
        qld = next(row for row in fields["state_rows"] if row["state_code"] == "QLD")
        self.assertEqual(wa["title_records"]["active"], 2)
        self.assertEqual(wa["title_records"]["pending_application"], 1)
        self.assertEqual(wa["well_records"]["total_layer_records"], 5)
        self.assertEqual(wa["well_records"]["known_petroleum_type_records"], 5)
        self.assertEqual(qld["well_records"]["total_layer_records"], 7)
        self.assertEqual(fields["external_or_unallocated"]["title_records"]["Territory of Ashmore and Cartier Islands"], 3)
        self.assertEqual(fields["external_or_unallocated"]["well_records"]["Outside of Australia"], 1)
        self.assertIn("Greenhouse Gas Assessment Permit", calls[0][1]["where"])

    def test_nopta_production_licence_map_keeps_volume_unavailable(self):
        def fake_get(_url, params=None, **_kwargs):
            self.assertIn("Production Licence", params["where"])
            return FakeResponse(
                json_doc={
                    "features": [
                        {
                            "attributes": {
                                "Title": "WA-1-L",
                                "RelTitle": "WA-1-P",
                                "TitleType": "Production Licence",
                                "ExpiryDate": None,
                                "GrantDate": 0,
                                "Status": "Active",
                                "FieldName": "Example Field",
                                "BasinName": "Northern Carnarvon Basin",
                                "SubBasin": None,
                                "OffShoreAr": "Western Australia",
                                "TitleOprat": "Example Operator Pty Ltd",
                                "TitleHold": "Example Operator Pty Ltd, Example Partner Pty Ltd",
                                "NoOfBlocks": 2,
                                "AreaKM2": 123.456,
                                "NEATS_Links": "https://example.test/title",
                            }
                        }
                    ]
                }
            )

        source = {
            "id": "state_petroleum_production_licence_map",
            "fetch_url": "https://example.test/TitlesCompany/query",
        }
        with mock.patch.object(fetch_data.requests, "get", side_effect=fake_get):
            out = fetch_data.fetch_nopta_production_licence_map(source)

        self.assertEqual(out["unit"], "active offshore petroleum production licence records")
        self.assertEqual(out["values"][0]["v"], 1)
        fields = out["extra"]["fields"]
        wa = next(row for row in fields["state_rows"] if row["state_code"] == "WA")
        self.assertEqual(wa["production_licence_records"], 1)
        mapped = fields["production_licence_rows"][0]
        self.assertEqual(mapped["field_name"], "Example Field")
        self.assertEqual(mapped["title_operator"], "Example Operator Pty Ltd")
        self.assertIsNone(mapped["production_metric_value"])
        self.assertIn("not a production-volume", mapped["notes"])

    def test_remp_oil_gas_projects_marks_capacity_not_production(self):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Oil & gas"
        ws.append(["intro"])
        ws.append(["Project", "Company", "State", "Latitude", "Longitude", "Type", "Status", "Annual Estimated New Capacity", "Capacity Unit", "Resource", "Construction Employment Estimate", "Operating Employment Estimate", "Cost Estimate A$m", "Estimated Start Commercial Operation"])
        ws.append(["Gas Project", "Example Energy", "QLD", -27.0, 150.0, "New project", "Committed", "82", "TJ/d", "Gas", 1, 2, "500", "2026"])

        source = {
            "id": "state_oil_gas_major_projects_remp",
            "fetch_url": "https://example.test/remp.xlsx",
            "remp_sheet": "Oil & gas",
        }
        with mock.patch.object(fetch_data, "load_xlsx_from_url", return_value=wb):
            out = fetch_data.fetch_remp_oil_gas_projects(source)

        self.assertEqual(out["unit"], "oil and gas major project rows")
        self.assertEqual(out["values"], [{"t": "2024-12-20", "v": 1}])
        fields = out["extra"]["fields"]
        self.assertEqual(fields["state_rows"][1]["project_rows"], 1)
        project = fields["project_rows"][0]
        self.assertEqual(project["project_name"], "Gas Project")
        self.assertEqual(project["company_name"], "Example Energy")
        self.assertEqual(project["production_metric_value"], 82.0)
        self.assertIn("not current production", project["production_metric_name"])
        self.assertIn("current production volume", project["notes"])


if __name__ == "__main__":
    unittest.main()
