#!/usr/bin/env python3
"""
fetch_data.py — Fuel Resilience AU data pipeline

Reads data/sources.yml, walks each source, and produces a normalised JSON file
under data/generated/<id>.json for every entry whose "fetch" key is
"programmatic". Files for "manual" or "unavailable" sources are left alone.

Design rules
------------
* Standard library + requests + PyYAML, plus openpyxl for selected publisher
  XLSX workbooks. No heavy data-frame dependencies.
* Never fabricate values. If a programmatic fetch fails, exit non-zero.
* Blocking URL checks apply only to programmatic fetch_url endpoints.
* Broad source landing-page link health is available as a separate, non-blocking
  report mode for CI/workflow logs.
"""

from __future__ import annotations

import argparse
import calendar
import csv
import datetime as dt
import io
import json
import os
import pathlib
import re
import sys
import urllib.parse
import xml.etree.ElementTree as ET
from typing import Any, Callable

try:
    import yaml
except ImportError:
    sys.stderr.write("PyYAML is required. Install with: pip install pyyaml\n")
    sys.exit(2)

try:
    import requests
except ImportError:
    sys.stderr.write("requests is required. Install with: pip install requests\n")
    sys.exit(2)

ROOT = pathlib.Path(__file__).resolve().parent.parent
SOURCES_FILE = ROOT / "data" / "sources.yml"
GENERATED_DIR = ROOT / "data" / "generated"
MANUAL_DIR = ROOT / "data" / "manual"

UA = "FuelResilienceAU-DataBot/1.0 (+https://github.com/WowCorey/fuel-fertilizer-dashboard)"


def fetch_eia_series(series_code: str, url: str | None = None) -> dict[str, Any]:
    """Fetch a daily EIA crude spot-price series via the FRED CSV mirror."""
    fred_id = {"RBRTE": "DCOILBRENTEU", "RWTC": "DCOILWTICO"}.get(series_code)
    if not fred_id:
        raise RuntimeError(f"Unknown EIA series code: {series_code}")

    url = url or f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={fred_id}"
    r = requests.get(url, headers={"User-Agent": UA}, timeout=30)
    r.raise_for_status()

    reader = csv.reader(io.StringIO(r.text))
    try:
        header = next(reader)
    except StopIteration as exc:
        raise RuntimeError("FRED CSV was empty") from exc

    normalized_header = [h.strip().lower() for h in header]
    accepted_headers = [["date", fred_id.lower()], ["observation_date", fred_id.lower()]]
    if normalized_header not in accepted_headers:
        raise RuntimeError(f"Unexpected FRED CSV header: {header!r}")

    by_month: dict[str, list[float]] = {}
    for row in reader:
        if len(row) < 2:
            continue
        date_str, val_str = row[0], row[1]
        if val_str in ("", "."):
            continue
        try:
            v = float(val_str)
        except ValueError:
            continue
        month_key = date_str[:7]
        by_month.setdefault(month_key, []).append(v)

    values = []
    for month in sorted(by_month):
        mean_v = sum(by_month[month]) / len(by_month[month])
        values.append({"t": month, "v": round(mean_v, 2)})

    values = values[-60:]
    if not values:
        raise RuntimeError(f"No numeric observations found in {fred_id}")

    last = values[-1]["t"]
    return {
        "unit": "USD per barrel",
        "values": values,
        "last_data_point": f"{last}-01",
        "notes": (
            f"Monthly mean of daily FRED series {fred_id} "
            f"(mirror of EIA {series_code}). Trimmed to last 60 months."
        ),
        "source_url_resolved": url,
    }


def fetch_fred_fuel_csv(source: dict[str, Any]) -> dict[str, Any]:
    """Fetch a FRED CSV mirror of an EIA refined-fuel price series.

    Shared by weekly retail and daily spot series. Values are aggregated to
    monthly means; last_data_point reflects the latest raw observation so the
    envelope's freshness matches the publisher's cadence.
    """
    url = source.get("fetch_url")
    sid = source.get("id", "<unknown>")
    if not url:
        raise RuntimeError(f"{sid}: fetch_url is required")

    parsed_qs = urllib.parse.parse_qs(urllib.parse.urlparse(url).query)
    fred_id = (source.get("fred_series") or "").strip() or (parsed_qs.get("id", [""])[0]).strip()
    if not fred_id:
        raise RuntimeError(f"{sid}: cannot determine FRED series id from fetch_url or fred_series")
    unit = source.get("fetch_unit") or "USD per gallon"

    r = requests.get(url, headers={"User-Agent": UA}, timeout=45)
    r.raise_for_status()

    reader = csv.reader(io.StringIO(r.text))
    try:
        header = next(reader)
    except StopIteration as exc:
        raise RuntimeError(f"FRED CSV for {fred_id} was empty") from exc

    normalized = [h.strip().lower() for h in header]
    accepted = {("observation_date", fred_id.lower()), ("date", fred_id.lower())}
    if tuple(normalized) not in accepted:
        raise RuntimeError(f"Unexpected FRED CSV header for {fred_id}: {header!r}")

    by_month: dict[str, list[float]] = {}
    latest_raw: str | None = None
    for row in reader:
        if len(row) < 2:
            continue
        date_str, val_str = row[0].strip(), row[1].strip()
        if val_str in ("", "."):
            continue
        try:
            v = float(val_str)
        except ValueError:
            continue
        if len(date_str) < 10:
            continue
        by_month.setdefault(date_str[:7], []).append(v)
        if latest_raw is None or date_str > latest_raw:
            latest_raw = date_str

    values = [
        {"t": month, "v": round(sum(vs) / len(vs), 3)}
        for month, vs in sorted(by_month.items())
    ][-60:]

    if not values or latest_raw is None:
        raise RuntimeError(f"No numeric observations found in {fred_id}")

    return {
        "unit": unit,
        "values": values,
        "last_data_point": latest_raw,
        "notes": (
            f"Monthly mean of FRED series {fred_id} (mirror of EIA). "
            "Trimmed to last 60 months; last_data_point reflects the latest raw observation."
        ),
        "source_url_resolved": url,
    }


def fetch_abs_sdmx(
    dataflow: str,
    key: str,
    start_period: str | None = None,
    classification_note: str | None = None,
    fetch_url: str | None = None,
) -> dict[str, Any]:
    """Fetch one ABS SDMX-JSON time series from the ABS Data API."""
    if not dataflow or not key:
        raise RuntimeError("ABS SDMX fetch requires dataflow and key")

    if fetch_url:
        url = fetch_url
    else:
        query = {"format": "jsondata"}
        if start_period:
            query["startPeriod"] = start_period
        url = f"https://data.api.abs.gov.au/rest/data/{dataflow}/{key}?{urllib.parse.urlencode(query)}"

    r = requests.get(
        url,
        headers={
            "User-Agent": UA,
            "Accept": "application/vnd.sdmx.data+json",
        },
        timeout=45,
    )
    r.raise_for_status()

    try:
        doc = r.json()
    except ValueError as exc:
        raise RuntimeError("ABS SDMX response was not valid JSON") from exc

    data_node = doc.get("data", doc)
    data_sets = data_node.get("dataSets")
    structures = data_node.get("structures")
    structure = doc.get("structure")
    if structure is None and isinstance(structures, list) and structures:
        structure = structures[0]

    if not isinstance(data_sets, list) or not data_sets:
        raise RuntimeError("ABS SDMX response contained no dataSets")
    if not isinstance(structure, dict):
        raise RuntimeError("ABS SDMX response contained no usable structure")

    data_set = data_sets[0]
    series = data_set.get("series")
    if not isinstance(series, dict) or not series:
        raise RuntimeError("ABS SDMX response contained no series")
    non_empty_series = [
        item for item in series.values()
        if isinstance(item, dict) and item.get("observations")
    ]
    if not non_empty_series:
        raise RuntimeError("ABS SDMX response contained no observations")
    if len(non_empty_series) != 1:
        raise RuntimeError(f"ABS SDMX key returned {len(non_empty_series)} series; expected exactly one")

    dimensions = structure.get("dimensions")
    if not isinstance(dimensions, dict):
        raise RuntimeError("ABS SDMX structure contained no dimensions")
    obs_dimensions = dimensions.get("observation")
    if not isinstance(obs_dimensions, list) or not obs_dimensions:
        raise RuntimeError("ABS SDMX structure contained no observation dimensions")
    time_values = obs_dimensions[0].get("values")
    if not isinstance(time_values, list) or not time_values:
        raise RuntimeError("ABS SDMX structure contained no time values")

    values: list[dict[str, Any]] = []
    period_ends: dict[str, str] = {}
    observations = non_empty_series[0].get("observations", {})
    for obs_idx, obs in observations.items():
        try:
            time_value = time_values[int(obs_idx)]
        except (ValueError, IndexError) as exc:
            raise RuntimeError(f"ABS SDMX observation index {obs_idx!r} has no matching time period") from exc
        if not isinstance(obs, list) or not obs:
            continue
        raw_value = obs[0]
        if raw_value is None:
            continue
        if isinstance(raw_value, bool) or not isinstance(raw_value, (int, float)):
            raise RuntimeError(f"ABS SDMX observation for {time_value.get('id')} was not numeric")
        period = time_value.get("id") or time_value.get("name")
        if not isinstance(period, str) or not period:
            raise RuntimeError("ABS SDMX time value was missing an id")
        period_end = time_value.get("end")
        if isinstance(period_end, str) and len(period_end) >= 10:
            period_ends[period] = period_end[:10]
        values.append({"t": period, "v": raw_value})

    values = sorted(values, key=lambda point: point["t"])[-60:]
    if not values:
        raise RuntimeError("ABS SDMX response contained no numeric observations")

    def attr_label(attr_id: str) -> str | None:
        attrs = structure.get("attributes", {})
        if not isinstance(attrs, dict):
            return None
        for group in ("dataSet", "series", "observation"):
            for attr in attrs.get(group, []) or []:
                if attr.get("id") != attr_id:
                    continue
                attr_values = attr.get("values") or []
                if not attr_values:
                    return None
                value_id = str(attr_values[0].get("id", "")).strip()
                value_name = str(attr_values[0].get("name", "")).strip()
                return value_id or value_name or None
        return None

    unit_measure = attr_label("UNIT_MEASURE") or "AUD"
    unit_mult = attr_label("UNIT_MULT")
    unit = unit_measure
    if unit_measure == "AUD" and unit_mult == "3":
        unit = "AUD thousands"
    elif unit_mult:
        unit = f"{unit_measure} (unit multiplier {unit_mult})"

    last = values[-1]["t"]
    last_data_point = period_ends.get(last) or (f"{last}-01" if len(last) == 7 else last)
    notes = (
        f"Fetched from ABS Data API dataflow {dataflow}, key {key}. "
        "Values are sorted by TIME_PERIOD and trimmed to the last 60 monthly observations."
    )
    if classification_note:
        notes = f"{notes} {classification_note}"
    return {
        "unit": unit,
        "values": values,
        "last_data_point": last_data_point,
        "notes": notes,
        "source_url_resolved": url,
    }


def fetch_abs_fertiliser_source_concentration(source: dict[str, Any]) -> dict[str, Any]:
    """Fetch ABS SITC 562 import values by source country and compute top-3 share."""
    dataflow = source.get("fetch_dataflow")
    key = source.get("fetch_key")
    if not dataflow or not key:
        raise RuntimeError(f"{source['id']}: fetch_dataflow and fetch_key are required")

    if source.get("fetch_url"):
        url = source["fetch_url"]
    else:
        query = {"format": "jsondata"}
        if source.get("fetch_start_period"):
            query["startPeriod"] = source["fetch_start_period"]
        url = f"https://data.api.abs.gov.au/rest/data/{dataflow}/{key}?{urllib.parse.urlencode(query)}"

    r = requests.get(
        url,
        headers={
            "User-Agent": UA,
            "Accept": "application/vnd.sdmx.data+json",
        },
        timeout=60,
    )
    r.raise_for_status()
    try:
        doc = r.json()
    except ValueError as exc:
        raise RuntimeError("ABS source-country response was not valid JSON") from exc

    data_node = doc.get("data", doc)
    data_sets = data_node.get("dataSets")
    structures = data_node.get("structures")
    structure = doc.get("structure")
    if structure is None and isinstance(structures, list) and structures:
        structure = structures[0]
    if not isinstance(data_sets, list) or not data_sets:
        raise RuntimeError("ABS source-country response contained no dataSets")
    if not isinstance(structure, dict):
        raise RuntimeError("ABS source-country response contained no usable structure")

    dimensions = structure.get("dimensions")
    if not isinstance(dimensions, dict):
        raise RuntimeError("ABS source-country structure contained no dimensions")
    series_dims = dimensions.get("series")
    obs_dims = dimensions.get("observation")
    if not isinstance(series_dims, list) or not isinstance(obs_dims, list) or not obs_dims:
        raise RuntimeError("ABS source-country structure was missing series or observation dimensions")

    dim_positions = {dim.get("id"): idx for idx, dim in enumerate(series_dims)}
    required_dims = ("COMMODITY_SITC", "COUNTRY_ORIGIN", "STATE_DEST", "FREQ")
    missing_dims = [dim for dim in required_dims if dim not in dim_positions]
    if missing_dims:
        raise RuntimeError(f"ABS source-country structure missing dimensions: {', '.join(missing_dims)}")

    time_values = obs_dims[0].get("values")
    if not isinstance(time_values, list) or not time_values:
        raise RuntimeError("ABS source-country structure contained no time values")

    values_by_period: dict[str, dict[str, float]] = {}
    totals_by_period: dict[str, float] = {}
    period_ends: dict[str, str] = {}
    target_commodity = source.get("fetch_commodity", "562")

    data_set = data_sets[0]
    series = data_set.get("series")
    if not isinstance(series, dict) or not series:
        raise RuntimeError("ABS source-country response contained no series")

    for series_key, series_doc in series.items():
        if not isinstance(series_doc, dict) or not series_doc.get("observations"):
            continue
        try:
            key_indices = [int(part) for part in series_key.split(":")]
        except ValueError as exc:
            raise RuntimeError(f"ABS source-country series key was malformed: {series_key!r}") from exc

        def dim_value(dim_id: str) -> dict[str, Any]:
            pos = dim_positions[dim_id]
            try:
                dim = series_dims[pos]
                return dim["values"][key_indices[pos]]
            except (IndexError, KeyError, TypeError) as exc:
                raise RuntimeError(f"ABS source-country key {series_key!r} did not match dimension {dim_id}") from exc

        commodity = dim_value("COMMODITY_SITC")
        country = dim_value("COUNTRY_ORIGIN")
        state = dim_value("STATE_DEST")
        freq = dim_value("FREQ")
        if commodity.get("id") != target_commodity or state.get("id") != "TOT" or freq.get("id") != "M":
            continue

        country_id = country.get("id")
        country_name = country.get("name") or country_id
        if not isinstance(country_name, str) or not country_name:
            raise RuntimeError("ABS source-country series had no country name")

        for obs_idx, obs in series_doc.get("observations", {}).items():
            try:
                time_value = time_values[int(obs_idx)]
            except (ValueError, IndexError) as exc:
                raise RuntimeError(f"ABS source-country observation index {obs_idx!r} has no time period") from exc
            if not isinstance(obs, list) or not obs:
                continue
            raw_value = obs[0]
            if raw_value is None:
                continue
            if isinstance(raw_value, bool) or not isinstance(raw_value, (int, float)):
                raise RuntimeError(f"ABS source-country observation for {country_name} was not numeric")
            period = time_value.get("id") or time_value.get("name")
            if not isinstance(period, str) or not period:
                raise RuntimeError("ABS source-country time value was missing an id")
            period_end = time_value.get("end")
            if isinstance(period_end, str) and len(period_end) >= 10:
                period_ends[period] = period_end[:10]
            if country_id == "TOT":
                totals_by_period[period] = float(raw_value)
            else:
                values_by_period.setdefault(period, {})[country_name] = float(raw_value)

    values: list[dict[str, Any]] = []
    latest_fields: dict[str, Any] | None = None
    for period in sorted(values_by_period):
        country_values = {country: value for country, value in values_by_period[period].items() if value > 0}
        total = totals_by_period.get(period) or sum(country_values.values())
        if total <= 0 or not country_values:
            continue
        ranked = sorted(country_values.items(), key=lambda item: item[1], reverse=True)
        top3 = ranked[:3]
        top3_total = sum(value for _, value in top3)
        top3_share = round(top3_total / total * 100, 1)
        values.append({"t": period, "v": top3_share})
        hhi = round(sum((value / total * 100) ** 2 for value in country_values.values()), 1)
        latest_fields = {
            "commodity": "SITC 562 manufactured fertilisers",
            "period": period,
            "total_import_value_aud_thousands": round(total),
            "top3_share_percent": top3_share,
            "top3_value_aud_thousands": round(top3_total),
            "positive_country_count": len(country_values),
            "hhi": hhi,
            "top_countries": [
                {
                    "country": country,
                    "value_aud_thousands": round(value),
                    "share_percent": round(value / total * 100, 1),
                }
                for country, value in top3
            ],
        }

    values = values[-60:]
    if not values or latest_fields is None:
        raise RuntimeError("ABS source-country response contained no usable SITC 562 country observations")

    last = values[-1]["t"]
    return {
        "unit": "%",
        "values": values,
        "last_data_point": period_ends.get(last) or (f"{last}-01" if len(last) == 7 else last),
        "notes": (
            "Fetched from ABS Data API MERCH_IMP using SITC 562 by country of origin. "
            "Each value is the share of monthly SITC 562 import value from the three largest "
            "non-total source countries. Values are trimmed to the last 60 monthly observations."
        ),
        "source_url_resolved": url,
        "extra": {
            "schema": "abs_sitc562_source_concentration.v1",
            "fields": latest_fields,
        },
    }


def fetch_abs_petroleum_imports_yoy(source: dict[str, Any]) -> dict[str, Any]:
    """Derive YoY percentage change from the generated ABS petroleum imports envelope."""
    parent_id = source.get("derived_from") or "abs_petroleum_imports"
    parent_path = GENERATED_DIR / f"{parent_id}.json"
    if not parent_path.exists():
        raise RuntimeError(f"Cannot derive {source['id']}; missing {parent_path.relative_to(ROOT)}")

    with parent_path.open("r", encoding="utf-8") as f:
        parent = json.load(f)

    if parent.get("status") != "ok":
        raise RuntimeError(f"Cannot derive {source['id']}; parent {parent_id} status is not ok")
    parent_values = parent.get("values")
    if not isinstance(parent_values, list) or len(parent_values) < 13:
        raise RuntimeError(f"Cannot derive {source['id']}; parent {parent_id} has fewer than 13 observations")

    by_month: dict[str, float] = {}
    for point in parent_values:
        if not isinstance(point, dict):
            continue
        t = point.get("t")
        v = point.get("v")
        if isinstance(t, str) and isinstance(v, (int, float)) and not isinstance(v, bool):
            by_month[t] = float(v)

    values: list[dict[str, Any]] = []
    for month in sorted(by_month):
        try:
            current_month = dt.date.fromisoformat(f"{month}-01")
        except ValueError:
            continue
        prior_year = current_month.year - 1
        prior_month = f"{prior_year:04d}-{current_month.month:02d}"
        previous = by_month.get(prior_month)
        if previous in (None, 0):
            continue
        yoy = (by_month[month] - previous) / previous * 100
        values.append({"t": month, "v": round(yoy, 1)})

    if not values:
        raise RuntimeError(f"Cannot derive {source['id']}; no comparable year-on-year observations")

    return {
        "unit": "%",
        "values": values[-60:],
        "last_data_point": parent.get("last_data_point"),
        "notes": (
            "Derived from abs_petroleum_imports by comparing each month to the same month one year earlier. "
            "Source: ABS International Merchandise Trade."
        ),
        "source_url_resolved": parent.get("source_url"),
    }


def load_parent_envelope(parent_id: str) -> dict[str, Any]:
    """Load a generated or manual parent envelope by source id."""
    for base in (GENERATED_DIR, MANUAL_DIR):
        path = base / f"{parent_id}.json"
        if path.exists():
            with path.open("r", encoding="utf-8") as f:
                parent = json.load(f)
            if not isinstance(parent, dict):
                raise RuntimeError(f"Parent envelope {path.relative_to(ROOT)} is not a JSON object")
            return parent
    raise RuntimeError(f"Missing parent envelope for {parent_id}")


def fetch_extra_field_derived(source: dict[str, Any]) -> dict[str, Any]:
    """Derive a single-point series by selecting a typed extra.fields value.

    This is used for fuel-security dashboard slots where the source publishes
    several product-specific values inside one public table. The parent remains
    the auditable source envelope; this generated envelope gives each product a
    first-class source id for cards, coverage and freshness display.
    """
    parent_id = source.get("derived_from")
    field = source.get("derived_field")
    if not parent_id or not field:
        raise RuntimeError(f"{source.get('id', '<unknown>')}: derived_from and derived_field are required")

    parent = load_parent_envelope(parent_id)
    if parent.get("status") != "ok":
        raise RuntimeError(f"Cannot derive {source['id']}; parent {parent_id} status is not ok")
    fields = parent.get("extra", {}).get("fields", {})
    if not isinstance(fields, dict):
        raise RuntimeError(f"Cannot derive {source['id']}; parent {parent_id} has no typed extra.fields")
    value = fields.get(field)
    if not isinstance(value, (int, float)):
        raise RuntimeError(f"Cannot derive {source['id']}; parent field {field!r} is missing or not numeric")

    point_date = fields.get("as_at") or parent.get("last_data_point")
    if not isinstance(point_date, str) or not point_date:
        raise RuntimeError(f"Cannot derive {source['id']}; parent {parent_id} has no date")

    label = source.get("derived_label") or field
    return {
        "unit": source.get("fetch_unit") or parent.get("unit") or "",
        "values": [{"t": point_date, "v": round(float(value), 1)}],
        "last_data_point": parent.get("last_data_point"),
        "notes": (
            f"Derived by selecting {field} from {parent_id}. "
            f"This is a product-specific reshaping of the PM&C/DCCEEW public table, not an independent estimate."
        ),
        "extra": {
            "schema": "fuel_security_product_metric.v1",
            "fields": {
                "product": label,
                "parent_source_id": parent_id,
                "parent_field": field,
                "parent_last_data_point": parent.get("last_data_point"),
                "parent_manual_entry": bool(parent.get("manual_entry")),
                "method": "typed_extra_field_selection",
            },
        },
        "source_url_resolved": parent.get("source_url"),
    }


def fetch_rba_f11(url: str) -> dict[str, Any]:
    """Fetch RBA Table F11.1 CSV and return monthly mean AUD/USD values."""
    r = requests.get(url, headers={"User-Agent": UA}, timeout=45)
    r.raise_for_status()

    reader = list(csv.reader(io.StringIO(r.content.decode("utf-8-sig"))))
    if len(reader) < 12:
        raise RuntimeError("RBA F11.1 CSV was too short")

    header_idx = None
    usd_col = None
    for idx, row in enumerate(reader):
        if row and row[0] == "Title":
            header_idx = idx
            try:
                usd_col = row.index("A$1=USD")
            except ValueError as exc:
                raise RuntimeError("RBA F11.1 CSV did not contain A$1=USD column") from exc
            break
    if header_idx is None or usd_col is None:
        raise RuntimeError("RBA F11.1 CSV did not contain a Title header row")

    data_start = None
    for idx, row in enumerate(reader[header_idx + 1 :], start=header_idx + 1):
        if row and row[0] == "Series ID":
            data_start = idx + 1
            break
    if data_start is None:
        raise RuntimeError("RBA F11.1 CSV did not contain a Series ID row")

    by_month: dict[str, list[float]] = {}
    latest_date: dt.date | None = None
    for row in reader[data_start:]:
        if len(row) <= usd_col or not row or not row[0].strip():
            continue
        try:
            obs_date = dt.datetime.strptime(row[0].strip(), "%d-%b-%Y").date()
            value = float(row[usd_col])
        except (ValueError, TypeError):
            continue
        latest_date = max(latest_date, obs_date) if latest_date else obs_date
        by_month.setdefault(obs_date.strftime("%Y-%m"), []).append(value)

    values = [
        {"t": month, "v": round(sum(month_values) / len(month_values), 4)}
        for month, month_values in sorted(by_month.items())
        if month_values
    ][-60:]
    if not values or latest_date is None:
        raise RuntimeError("RBA F11.1 CSV contained no numeric AUD/USD observations")

    return {
        "unit": "USD per AUD",
        "values": values,
        "last_data_point": latest_date.isoformat(),
        "notes": "Monthly mean of daily AUD/USD observations from RBA Table F11.1. Trimmed to last 60 months.",
        "source_url_resolved": url,
    }


def fetch_rba_csv_column(
    url: str,
    title_substring: str,
    *,
    aggregate: str = "monthly_mean",
    unit: str,
    notes: str,
    ndigits: int = 4,
) -> dict[str, Any]:
    """Generic RBA Statistical Table CSV fetcher.

    RBA tables share a layout: a metadata block (rows starting with "Title",
    "Description", "Frequency", ..., "Series ID"), then date+value rows where
    column 0 is the observation date in DD-MMM-YYYY format. We pick the column
    whose Title row contains ``title_substring`` (case-insensitive substring
    match) and aggregate either to monthly mean or quarter-end.
    """
    if aggregate not in {"monthly_mean", "quarter_end"}:
        raise RuntimeError(f"fetch_rba_csv_column: unsupported aggregate {aggregate!r}")

    r = requests.get(url, headers={"User-Agent": UA}, timeout=45)
    r.raise_for_status()
    reader = list(csv.reader(io.StringIO(r.content.decode("utf-8-sig"))))
    if len(reader) < 12:
        raise RuntimeError(f"RBA CSV {url} was too short")

    target = title_substring.strip().lower()
    title_row_idx = None
    col_idx = None
    for idx, row in enumerate(reader):
        if not row or row[0] != "Title":
            continue
        title_row_idx = idx
        for ci, cell in enumerate(row[1:], start=1):
            if cell and target in cell.strip().lower():
                col_idx = ci
                break
        if col_idx is not None:
            break
    if title_row_idx is None or col_idx is None:
        raise RuntimeError(f"RBA CSV {url} did not contain a Title column matching {title_substring!r}")

    data_start = None
    for idx, row in enumerate(reader[title_row_idx + 1 :], start=title_row_idx + 1):
        if row and row[0] == "Series ID":
            data_start = idx + 1
            break
    if data_start is None:
        raise RuntimeError(f"RBA CSV {url} did not contain a Series ID row")

    by_period: dict[str, list[float]] = {}
    period_end_dates: dict[str, dt.date] = {}
    latest_date: dt.date | None = None
    for row in reader[data_start:]:
        if not row or len(row) <= col_idx or not row[0].strip():
            continue
        try:
            obs_date = dt.datetime.strptime(row[0].strip(), "%d-%b-%Y").date()
            value = float(row[col_idx])
        except (ValueError, TypeError):
            continue
        latest_date = max(latest_date, obs_date) if latest_date else obs_date
        if aggregate == "monthly_mean":
            key = obs_date.strftime("%Y-%m")
        else:
            quarter = (obs_date.month - 1) // 3 + 1
            key = f"{obs_date.year}-Q{quarter}"
        by_period.setdefault(key, []).append(value)
        prior = period_end_dates.get(key)
        if prior is None or obs_date > prior:
            period_end_dates[key] = obs_date

    if not by_period or latest_date is None:
        raise RuntimeError(f"RBA CSV {url} contained no numeric observations for column {title_substring!r}")

    if aggregate == "monthly_mean":
        values = [
            {"t": period, "v": round(sum(obs) / len(obs), ndigits)}
            for period, obs in sorted(by_period.items())
        ][-60:]
    else:
        values = [
            {"t": period, "v": round(obs[-1], ndigits)}
            for period, obs in sorted(by_period.items())
        ][-40:]

    return {
        "unit": unit,
        "values": values,
        "last_data_point": latest_date.isoformat(),
        "notes": notes,
        "source_url_resolved": url,
    }


def fetch_aemo_nem_price_demand(metric: str) -> dict[str, Any]:
    """Fetch AEMO NEM monthly price/demand CSVs across all 5 regions.

    Pulls the static CSVs at
    aemo.com.au/aemo/data/nem/priceanddemand/PRICE_AND_DEMAND_YYYYMM_REGION.csv
    for the last 12 complete months across NSW1, VIC1, QLD1, SA1, TAS1, then
    reduces 5-minute TRADE intervals to monthly means per region. The NEM-wide
    series is the simple mean of region means for ``metric == "price"`` and the
    sum of region means for ``metric == "demand"``. Per-region latest-month
    values land in ``extra.fields``.
    """
    if metric not in {"price", "demand"}:
        raise RuntimeError(f"fetch_aemo_nem_price_demand: unsupported metric {metric!r}")

    regions = ["NSW1", "VIC1", "QLD1", "SA1", "TAS1"]
    today = dt.date.today()
    months: list[dt.date] = []
    cursor = dt.date(today.year, today.month, 1)
    for _ in range(13):
        cursor = cursor.replace(day=1) - dt.timedelta(days=1)
        cursor = cursor.replace(day=1)
        months.append(cursor)
    months.reverse()

    base = "https://aemo.com.au/aemo/data/nem/priceanddemand/PRICE_AND_DEMAND_{ym}_{region}.csv"
    region_monthly: dict[str, dict[str, float]] = {r: {} for r in regions}

    for month in months:
        ym = month.strftime("%Y%m")
        for region in regions:
            url = base.format(ym=ym, region=region)
            r = requests.get(url, headers={"User-Agent": UA}, timeout=60)
            if r.status_code == 404:
                continue
            r.raise_for_status()
            reader = csv.DictReader(io.StringIO(r.content.decode("utf-8-sig")))
            sum_v = 0.0
            n_v = 0
            for row in reader:
                if row.get("PERIODTYPE") != "TRADE":
                    continue
                try:
                    if metric == "price":
                        v = float(row["RRP"])
                    else:
                        v = float(row["TOTALDEMAND"])
                except (KeyError, ValueError, TypeError):
                    continue
                sum_v += v
                n_v += 1
            if n_v == 0:
                continue
            region_monthly[region][month.strftime("%Y-%m")] = sum_v / n_v

    all_months: set[str] = set()
    for region in regions:
        all_months.update(region_monthly[region].keys())
    if not all_months:
        raise RuntimeError("AEMO NEM price/demand fetch returned no monthly observations")

    sorted_months = sorted(all_months)
    values: list[dict[str, Any]] = []
    for ym in sorted_months:
        per_region = [region_monthly[r][ym] for r in regions if ym in region_monthly[r]]
        if not per_region:
            continue
        if metric == "price":
            agg = sum(per_region) / len(per_region)
            values.append({"t": ym, "v": round(agg, 2)})
        else:
            agg = sum(per_region)
            values.append({"t": ym, "v": round(agg, 0)})

    if not values:
        raise RuntimeError("AEMO NEM price/demand aggregation produced no values")

    latest_month = values[-1]["t"]
    region_latest: dict[str, Any] = {}
    for region in regions:
        v = region_monthly[region].get(latest_month)
        if v is None:
            region_latest[region] = None
        else:
            region_latest[region] = round(v, 2 if metric == "price" else 0)

    last_year, last_month = latest_month.split("-")
    last_day = calendar.monthrange(int(last_year), int(last_month))[1]
    last_data_point = f"{last_year}-{last_month}-{last_day:02d}"

    if metric == "price":
        unit = "AUD per MWh"
        notes = (
            "Monthly mean of 5-minute TRADE-interval Regional Reference Price across NSW1, VIC1, QLD1, SA1 and TAS1. "
            "Aggregate is the simple mean of the five regional monthly means. extra.fields holds the latest-month value per region."
        )
    else:
        unit = "MW"
        notes = (
            "Monthly mean of 5-minute TRADE-interval TOTALDEMAND across NSW1, VIC1, QLD1, SA1 and TAS1, then summed across regions. "
            "extra.fields holds the latest-month region-mean value per region."
        )

    return {
        "unit": unit,
        "values": values,
        "last_data_point": last_data_point,
        "notes": notes,
        "extra": {
            "schema": "fuel_resilience_typed_fields.v1",
            "fields": {
                "latest_month": latest_month,
                **{f"region_{r.lower()}_latest": region_latest[r] for r in regions},
            },
        },
        "source_url_resolved": base.format(ym=latest_month.replace("-", ""), region="NSW1"),
    }


def month_end_iso(month: dt.date) -> str:
    last_day = calendar.monthrange(month.year, month.month)[1]
    return dt.date(month.year, month.month, last_day).isoformat()


def load_xlsx_from_url(url: str, extra_headers: dict[str, str] | None = None):
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX fetchers. Install with: pip install openpyxl") from exc

    headers = {"User-Agent": UA}
    if extra_headers:
        headers.update(extra_headers)
    r = requests.get(url, headers=headers, timeout=90)
    r.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)


def normalized_abs_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\xa0", " ").replace("\u2019", "'")
    text = re.sub(r"\s+", " ", text).strip()
    return text.lstrip(">").strip().lower()


def discover_abs_release_xlsx_url(release_url: str, label_contains: str | list[str]) -> str:
    """Find a named XLSX download link on an ABS latest-release page."""
    needles = [label_contains] if isinstance(label_contains, str) else list(label_contains)
    wanted = [normalized_abs_text(needle) for needle in needles if normalized_abs_text(needle)]
    if not release_url:
        raise RuntimeError("ABS release URL is required")
    if release_url.lower().endswith(".xlsx"):
        return release_url
    if not wanted:
        raise RuntimeError("ABS XLSX discovery requires fetch_xlsx_label_contains")

    r = requests.get(release_url, headers={"User-Agent": UA}, timeout=60)
    r.raise_for_status()

    candidates: list[tuple[str, str]] = []
    for match in re.finditer(r"<a[^>]+href=[\"']([^\"']+\.xlsx)[\"'][^>]*>(.*?)</a>", r.text, flags=re.I | re.S):
        tag = match.group(0)
        href = urllib.parse.urljoin(r.url, match.group(1))
        aria = re.search(r"aria-label=[\"']([^\"']+)[\"']", tag, flags=re.I)
        label = aria.group(1) if aria else re.sub(r"<.*?>", " ", match.group(2), flags=re.S)
        label_norm = normalized_abs_text(label)
        candidates.append((label_norm, href))
        if all(needle in label_norm for needle in wanted):
            return href

    available = "; ".join(label for label, _ in candidates[:8])
    raise RuntimeError(f"ABS release page did not contain a matching XLSX link. Wanted {wanted}; saw {available}")


def _round_numeric(value: float, ndigits: int) -> int | float:
    rounded = round(float(value), ndigits)
    return int(rounded) if ndigits == 0 else rounded


def extract_abs_xlsx_series(
    wb: Any,
    targets: list[str],
    *,
    series_type: str | None,
    ndigits: int,
    limit: int,
) -> tuple[list[dict[str, Any]], str, list[dict[str, Any]], str]:
    """Extract one or more ABS spreadsheet series and sum them by period."""
    if not targets:
        raise RuntimeError("ABS XLSX extraction requires at least one target series description")

    wanted_type = normalized_abs_text(series_type) if series_type else None
    found: list[dict[str, Any]] = []

    for target in targets:
        target_norm = normalized_abs_text(target)
        match_info: dict[str, Any] | None = None
        for sheet_name in wb.sheetnames:
            if not str(sheet_name).lower().startswith("data"):
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 11:
                continue
            max_cols = max(len(row) for row in rows[:10])
            for col_idx in range(1, max_cols):
                desc = rows[0][col_idx] if col_idx < len(rows[0]) else None
                if not desc:
                    continue
                desc_norm = normalized_abs_text(desc)
                if target_norm not in desc_norm:
                    continue
                col_type = normalized_abs_text(rows[2][col_idx] if col_idx < len(rows[2]) else None)
                if wanted_type and col_type != wanted_type:
                    continue
                observations: dict[str, float] = {}
                for row in rows[10:]:
                    raw_date = row[0] if row else None
                    raw_value = row[col_idx] if col_idx < len(row) else None
                    if isinstance(raw_date, dt.datetime):
                        obs_date = raw_date.date()
                    elif isinstance(raw_date, dt.date):
                        obs_date = raw_date
                    else:
                        continue
                    if isinstance(raw_value, bool) or not isinstance(raw_value, (int, float)):
                        continue
                    observations[obs_date.strftime("%Y-%m")] = float(raw_value)
                if observations:
                    match_info = {
                        "target": target,
                        "description": str(desc),
                        "sheet": sheet_name,
                        "series_id": rows[9][col_idx] if col_idx < len(rows[9]) else None,
                        "unit": rows[1][col_idx] if col_idx < len(rows[1]) else None,
                        "series_type": rows[2][col_idx] if col_idx < len(rows[2]) else None,
                        "observations": observations,
                    }
                    break
            if match_info:
                break
        if not match_info:
            raise RuntimeError(f"ABS workbook did not contain target series {target!r}")
        found.append(match_info)

    common_periods = set(found[0]["observations"])
    for series in found[1:]:
        common_periods &= set(series["observations"])
    values: list[dict[str, Any]] = []
    for period in sorted(common_periods):
        total = sum(series["observations"][period] for series in found)
        values.append({"t": period, "v": _round_numeric(total, ndigits)})

    values = values[-limit:]
    if not values:
        raise RuntimeError("ABS workbook target series contained no common numeric observations")

    latest_period = values[-1]["t"]
    latest_date = dt.date.fromisoformat(f"{latest_period}-01")
    latest_components = [
        {
            "description": series["description"],
            "series_id": series["series_id"],
            "sheet": series["sheet"],
            "series_type": series["series_type"],
            "unit": series["unit"],
            "value": _round_numeric(series["observations"][latest_period], ndigits),
        }
        for series in found
    ]
    unit = str(found[0]["unit"] or "").strip()
    return values, month_end_iso(latest_date), latest_components, unit


def fetch_abs_release_xlsx_series(source: dict[str, Any]) -> dict[str, Any]:
    release_url = source.get("fetch_url") or source.get("canonical_url") or source.get("url")
    workbook_url = discover_abs_release_xlsx_url(release_url, source.get("fetch_xlsx_label_contains", []))
    wb = load_xlsx_from_url(workbook_url)

    targets = source.get("fetch_series_descriptions") or []
    if isinstance(targets, str):
        targets = [targets]
    values, last_data_point, components, workbook_unit = extract_abs_xlsx_series(
        wb,
        list(targets),
        series_type=source.get("fetch_series_type"),
        ndigits=int(source.get("fetch_round_digits", 1)),
        limit=int(source.get("fetch_trim_observations", 60)),
    )
    unit = source.get("fetch_unit") or workbook_unit
    component_names = "; ".join(str(component["series_id"]) for component in components)
    notes = source.get("fetch_notes") or (
        f"Fetched from ABS latest-release XLSX workbook. Series IDs: {component_names}. "
        f"Workbook unit: {workbook_unit}. Values are trimmed to the last {len(values)} observations."
    )
    return {
        "unit": unit,
        "values": values,
        "last_data_point": last_data_point,
        "notes": notes,
        "source_url_resolved": workbook_url,
        "extra": {
            "schema": "abs_xlsx_series.v1",
            "fields": {
                "workbook_url": workbook_url,
                "workbook_unit": workbook_unit,
                "component_series": components,
            },
        },
    }


def discover_aps_workbook_url(package_url: str) -> str:
    r = requests.get(package_url, headers={"User-Agent": UA}, timeout=45)
    r.raise_for_status()
    package = r.json().get("result", {})
    candidates = [
        res for res in package.get("resources", [])
        if str(res.get("format", "")).lower() in {"xlsx", "excel (.xlsx)"}
        and str(res.get("url", "")).lower().endswith(".xlsx")
    ]
    if not candidates:
        raise RuntimeError("APS CKAN package did not contain an XLSX resource")
    return candidates[0]["url"]


def sheet_series(
    wb: Any,
    sheet_name: str,
    column_name: str,
    ndigits: int = 1,
    limit: int | None = 60,
) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Workbook did not contain sheet {sheet_name!r}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError(f"Sheet {sheet_name!r} was empty")
    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    normalized = [h.replace("\n", " ").strip().lower() for h in header]
    target = column_name.replace("\n", " ").strip().lower()
    try:
        col_idx = normalized.index(target)
    except ValueError as exc:
        raise RuntimeError(f"Sheet {sheet_name!r} did not contain column {column_name!r}") from exc

    values: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not row or not isinstance(row[0], (dt.datetime, dt.date)):
            continue
        raw = row[col_idx] if col_idx < len(row) else None
        if isinstance(raw, bool) or not isinstance(raw, (int, float)):
            continue
        month = row[0].date() if isinstance(row[0], dt.datetime) else row[0]
        values.append({"t": month.strftime("%Y-%m"), "v": round(float(raw), ndigits)})
    if not values:
        raise RuntimeError(f"Sheet {sheet_name!r} column {column_name!r} contained no numeric observations")
    return values[-limit:] if limit else values


def sheet_dated_series(
    wb: Any,
    sheet_name: str,
    column_name: str,
    ndigits: int = 1,
) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Workbook did not contain sheet {sheet_name!r}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError(f"Sheet {sheet_name!r} was empty")
    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    normalized = [h.replace("\n", " ").strip().lower() for h in header]
    target = column_name.replace("\n", " ").strip().lower()
    try:
        col_idx = normalized.index(target)
    except ValueError as exc:
        raise RuntimeError(f"Sheet {sheet_name!r} did not contain column {column_name!r}") from exc

    values: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not row or not isinstance(row[0], (dt.datetime, dt.date)):
            continue
        raw = row[col_idx] if col_idx < len(row) else None
        if isinstance(raw, bool) or not isinstance(raw, (int, float)):
            continue
        obs_date = row[0].date() if isinstance(row[0], dt.datetime) else row[0]
        values.append({"t": obs_date.isoformat(), "v": round(float(raw), ndigits)})
    if not values:
        raise RuntimeError(f"Sheet {sheet_name!r} column {column_name!r} contained no numeric observations")
    return values


def fetch_aps_xlsx_series(source: dict[str, Any]) -> dict[str, Any]:
    package_url = source.get("fetch_url") or source.get("data_portal_url")
    if not package_url:
        raise RuntimeError(f"{source['id']}: APS fetch_url/data_portal_url is required")
    workbook_url = discover_aps_workbook_url(package_url)
    wb = load_xlsx_from_url(workbook_url)
    values = sheet_series(
        wb,
        source["aps_sheet"],
        source["aps_column"],
        int(source.get("fetch_round_digits", 1)),
    )
    last_month = dt.date.fromisoformat(values[-1]["t"] + "-01")
    return {
        "unit": source.get("fetch_unit", ""),
        "values": values,
        "last_data_point": month_end_iso(last_month),
        "notes": (
            f"Fetched from the Australian Petroleum Statistics XLSX data extract. "
            f"Sheet: {source['aps_sheet']}; column: {source['aps_column']}. "
            "Values are trimmed to the last 60 monthly observations."
        ),
        "source_url_resolved": workbook_url,
    }


def discover_aip_tgp_workbook_url(page_url: str) -> str:
    r = requests.get(page_url, headers={"User-Agent": UA}, timeout=45)
    r.raise_for_status()
    matches = re.findall(r"https?://[^\"']+/AIP_TGP_Data_[^\"']+\.xlsx", r.text)
    if not matches:
        raise RuntimeError("AIP TGP page did not link a dated AIP_TGP_Data XLSX workbook")
    return matches[-1]


def fetch_aip_tgp(source: dict[str, Any]) -> dict[str, Any]:
    page_url = source.get("fetch_url") or source.get("archive_url") or source.get("canonical_url")
    if not page_url:
        raise RuntimeError("aip_tgp requires a fetch_url/archive_url")
    workbook_url = discover_aip_tgp_workbook_url(page_url)
    wb = load_xlsx_from_url(workbook_url)
    raw_values = sheet_dated_series(wb, "Petrol TGP", "National Average", 1)

    by_month: dict[str, list[float]] = {}
    for point in raw_values:
        by_month.setdefault(point["t"][:7], []).append(point["v"])
    values = [
        {"t": month, "v": round(sum(month_values) / len(month_values), 1)}
        for month, month_values in sorted(by_month.items())
    ][-60:]
    if not values:
        raise RuntimeError("AIP TGP workbook contained no national average ULP observations")
    latest_raw_date = max(point["t"] for point in raw_values)
    return {
        "unit": "cents per litre",
        "values": values,
        "last_data_point": latest_raw_date,
        "notes": (
            "Monthly mean of daily national average unleaded petrol terminal gate prices "
            "from the AIP historical TGP XLSX workbook."
        ),
        "source_url_resolved": workbook_url,
    }


def warn_skip(label: str, message: str) -> None:
    print(f"[WARN] {label}: {message}", file=sys.stderr)


def retail_result(state: str, prices: list[float], source_date: str, detail: str) -> dict[str, Any] | None:
    prices = [p for p in prices if p > 0]
    if not prices:
        return None
    return {
        "state": state,
        "average": round(sum(prices) / len(prices), 1),
        "stations": len(prices),
        "date": source_date,
        "detail": detail,
    }


RETAIL_PRODUCTS: dict[str, dict[str, Any]] = {
    "ulp91": {
        "label": "ULP 91",
        "nsw_fuel_types": {"U91", "ULP", "UNLEADED", "REGULAR UNLEADED"},
        "qld_fuel_types": ["Unleaded"],
        "wa_product_code": 1,
    },
    "diesel": {
        "label": "diesel",
        "nsw_fuel_types": {"DL", "DIESEL"},
        "qld_fuel_types": ["Diesel"],
        "wa_product_code": 4,
    },
    "premium95": {
        "label": "premium 95",
        "nsw_fuel_types": {"P95", "PULP95", "PREMIUM 95", "PREMIUM UNLEADED 95"},
        "qld_fuel_types": ["PULP 95/96 RON"],
        "wa_product_code": 2,
    },
    "e10": {
        "label": "E10",
        "nsw_fuel_types": {"E10"},
        "qld_fuel_types": ["e10"],
        # WA FuelWatch's public RSS product-code list has no active E10 code.
        "wa_product_code": None,
    },
}


def fetch_wa_fuelwatch(url: str, label: str = "ULP 91") -> dict[str, Any] | None:
    try:
        r = requests.get(url, headers={"User-Agent": UA}, timeout=45)
        if r.status_code != 200:
            warn_skip("WA FuelWatch", f"HTTP {r.status_code}")
            return None
        root = ET.fromstring(r.content)
    except Exception as exc:
        warn_skip("WA FuelWatch", f"{type(exc).__name__}: {exc}")
        return None

    prices: list[float] = []
    dates: set[str] = set()
    for item in root.findall("./channel/item"):
        price_text = item.findtext("price")
        date_text = item.findtext("date")
        try:
            price = float(price_text) if price_text is not None else None
        except ValueError:
            continue
        if price is None:
            continue
        prices.append(price)
        if date_text:
            dates.add(date_text)

    source_date = sorted(dates)[-1] if dates else dt.datetime.now(dt.timezone.utc).date().isoformat()
    return retail_result("WA", prices, source_date, f"FuelWatch {label} RSS feed")


QLD_MONTH_NAMES = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}


def qld_resource_period(resource: dict[str, Any]) -> tuple[int, int] | None:
    text = f"{resource.get('name', '')} {resource.get('url', '')}".lower()
    year = None
    month = None
    if found := re.search(r"20\d{2}[-_](\d{2})", text):
        year_match = re.search(r"(20\d{2})[-_]\d{2}", text)
        if year_match:
            year = int(year_match.group(1))
            month = int(found.group(1))
    if year is None:
        if year_match := re.search(r"20\d{2}", text):
            year = int(year_match.group(0))
        for name, value in QLD_MONTH_NAMES.items():
            if name in text:
                month = value
                break
    if year and month:
        return year, month
    return None


def qld_datastore_resources(package_id: str, ckan_base: str) -> list[dict[str, Any]] | None:
    url = f"{ckan_base.rstrip('/')}/api/3/action/package_show?id={urllib.parse.quote(package_id)}"
    try:
        r = requests.get(url, headers={"User-Agent": UA}, timeout=45)
        if r.status_code != 200:
            warn_skip("QLD Fuel Prices", f"package_show HTTP {r.status_code}")
            return None
        package = r.json().get("result", {})
    except Exception as exc:
        warn_skip("QLD Fuel Prices", f"package lookup {type(exc).__name__}: {exc}")
        return None

    return [
        resource
        for resource in package.get("resources", [])
        if str(resource.get("format", "")).upper() == "CSV" and resource.get("datastore_active")
    ]


def latest_qld_resource(package_id: str, ckan_base: str) -> dict[str, Any] | None:
    resources = qld_datastore_resources(package_id, ckan_base)
    if resources is None:
        return None

    candidates: list[tuple[tuple[int, int], dict[str, Any]]] = []
    for resource in resources:
        period = qld_resource_period(resource)
        if period:
            candidates.append((period, resource))

    if not candidates:
        warn_skip("QLD Fuel Prices", "no datastore-backed monthly CSV resource found")
        return None
    return sorted(candidates, key=lambda item: item[0])[-1][1]


def fetch_qld_open_data(
    ckan_base: str,
    package_id: str,
    fuel_types: list[str] | None = None,
    label: str = "ULP 91",
) -> dict[str, Any] | None:
    resource = latest_qld_resource(package_id, ckan_base)
    if not resource:
        return None
    rid = resource.get("id")
    if not rid:
        warn_skip("QLD Fuel Prices", "latest resource did not have an id")
        return None

    fuel_types = fuel_types or ["Unleaded"]
    escaped_types = ", ".join("'" + fuel_type.replace("'", "''") + "'" for fuel_type in fuel_types)
    sql = (
        f'SELECT "SiteId", "Price", "TransactionDateutc" FROM "{rid}" '
        f'WHERE "Fuel_Type" IN ({escaped_types}) AND "Price" < 9999'
    )
    url = f"{ckan_base.rstrip('/')}/api/3/action/datastore_search_sql?{urllib.parse.urlencode({'sql': sql})}"
    try:
        r = requests.get(url, headers={"User-Agent": UA}, timeout=60)
        if r.status_code != 200:
            warn_skip("QLD Fuel Prices", f"datastore_search_sql HTTP {r.status_code}")
            return None
        records = r.json().get("result", {}).get("records", [])
    except Exception as exc:
        warn_skip("QLD Fuel Prices", f"datastore query {type(exc).__name__}: {exc}")
        return None

    latest_by_site: dict[str, tuple[str, float]] = {}
    for record in records:
        site_id = str(record.get("SiteId", "")).strip()
        date_text = str(record.get("TransactionDateutc", "")).strip()
        try:
            price = float(record.get("Price")) / 10
        except (TypeError, ValueError):
            continue
        if not site_id or not date_text:
            continue
        previous = latest_by_site.get(site_id)
        if previous is None or date_text > previous[0]:
            latest_by_site[site_id] = (date_text, price)

    if not latest_by_site:
        warn_skip("QLD Fuel Prices", f"latest public dataset contained no {label} prices")
        return None
    latest_date = max(date_text[:10] for date_text, _ in latest_by_site.values())
    detail = f"Queensland Open Data {label} latest public monthly file: {resource.get('name', rid)}"
    return retail_result("QLD", [price for _, price in latest_by_site.values()], latest_date, detail)


def qld_datastore_sql(ckan_base: str, sql: str) -> list[dict[str, Any]]:
    url = f"{ckan_base.rstrip('/')}/api/3/action/datastore_search_sql?{urllib.parse.urlencode({'sql': sql})}"
    r = requests.get(url, headers={"User-Agent": UA}, timeout=60)
    r.raise_for_status()
    doc = r.json()
    if not doc.get("success", True):
        raise RuntimeError(f"Queensland datastore query failed: {doc}")
    records = doc.get("result", {}).get("records", [])
    if not isinstance(records, list):
        raise RuntimeError("Queensland datastore query returned malformed records")
    return records


def fetch_qld_unavailable_reports(source: dict[str, Any]) -> dict[str, Any]:
    """Count QLD monthly fuel-price rows where Price = 9999.

    The Queensland Open Data column explanation defines 9999 as fuel stock
    temporarily unavailable. This is a partial monthly reporting signal, not a
    live station outage feed.
    """
    ckan_base = source.get("qld_ckan_base", "https://www.data.qld.gov.au")
    today = dt.datetime.now(dt.timezone(dt.timedelta(hours=10))).date()
    package_id = source.get("qld_package_id", f"fuel-price-reporting-{today.year}")
    resources = qld_datastore_resources(package_id, ckan_base)
    if not resources:
        raise RuntimeError(f"{source['id']}: no datastore-backed QLD monthly resources found")

    monthly: list[tuple[tuple[int, int], dict[str, Any], dict[str, Any]]] = []
    for resource in resources:
        period = qld_resource_period(resource)
        rid = resource.get("id")
        if not period or not rid:
            continue
        sql = (
            'SELECT COUNT(*) AS reports, COUNT(DISTINCT "SiteId") AS sites, '
            'MAX("TransactionDateutc") AS latest, MIN("TransactionDateutc") AS earliest '
            f'FROM "{rid}" WHERE "Price" = 9999'
        )
        records = qld_datastore_sql(ckan_base, sql)
        record = records[0] if records else {}
        monthly.append((period, resource, record))

    if not monthly:
        raise RuntimeError(f"{source['id']}: no dated QLD monthly resources could be queried")

    monthly = sorted(monthly, key=lambda item: item[0])[-60:]
    values = []
    for (year, month), _, record in monthly:
        try:
            reports = int(record.get("reports") or 0)
        except (TypeError, ValueError):
            raise RuntimeError(f"{source['id']}: malformed report count in QLD datastore response")
        values.append({"t": f"{year:04d}-{month:02d}", "v": reports})

    latest_period, latest_resource, latest_record = monthly[-1]
    latest_rid = latest_resource.get("id")
    latest_date = str(latest_record.get("latest") or "")[:10]
    if not latest_date:
        latest_date = month_end_iso(dt.date(latest_period[0], latest_period[1], 1))

    breakdown_records = []
    if latest_rid:
        sql = (
            'SELECT "Fuel_Type", COUNT(*) AS reports, COUNT(DISTINCT "SiteId") AS sites '
            f'FROM "{latest_rid}" WHERE "Price" = 9999 '
            'GROUP BY "Fuel_Type" ORDER BY reports DESC'
        )
        breakdown_records = qld_datastore_sql(ckan_base, sql)

    breakdown = []
    for record in breakdown_records:
        try:
            reports = int(record.get("reports") or 0)
            sites = int(record.get("sites") or 0)
        except (TypeError, ValueError):
            continue
        breakdown.append(
            {
                "fuel_type": str(record.get("Fuel_Type") or "Unknown"),
                "reports": reports,
                "sites": sites,
            }
        )

    try:
        latest_sites = int(latest_record.get("sites") or 0)
        latest_reports = int(latest_record.get("reports") or 0)
    except (TypeError, ValueError) as exc:
        raise RuntimeError(f"{source['id']}: malformed latest QLD stockout counts") from exc

    return {
        "unit": "unavailable fuel-type reports",
        "values": values,
        "last_data_point": latest_date,
        "notes": (
            "Counts rows in the latest monthly Queensland Open Data Fuel Price Reporting "
            "resource where Price = 9999. The official column explanation says 9999 "
            "denotes fuel stock temporarily unavailable, such as a tank empty and "
            "awaiting new stock. This is monthly change-report coverage, not a live "
            "statewide station outage count."
        ),
        "source_url_resolved": latest_resource.get("url") or source.get("url"),
        "extra": {
            "schema": "qld_fuel_security_unavailable_reports.v1",
            "fields": {
                "coverage": "Queensland monthly Open Data fuel-price reporting resources",
                "metric": "Rows where Price = 9999",
                "latest_resource_name": latest_resource.get("name"),
                "latest_resource_id": latest_rid,
                "latest_resource_url": latest_resource.get("url"),
                "earliest_unavailable_report_date": str(latest_record.get("earliest") or "")[:10] or None,
                "latest_unavailable_report_date": latest_date,
                "unavailable_fuel_type_reports": latest_reports,
                "distinct_sites_with_unavailable_fuel": latest_sites,
                "fuel_type_breakdown": breakdown,
                "not_covered": [
                    "live station-level outage status",
                    "national coverage",
                    "stations with no fuel of any product unless all products are reported unavailable",
                ],
            },
        },
    }


def fetch_nsw_fuelcheck(url: str, fuel_types: set[str] | None = None, label: str = "ULP 91") -> dict[str, Any] | None:
    token = os.environ.get("NSW_FUELCHECK_API_KEY", "").strip()
    if not token:
        warn_skip("NSW FuelCheck", "NSW_FUELCHECK_API_KEY is not set")
        return None
    try:
        r = requests.get(
            url,
            headers={
                "User-Agent": UA,
                "Authorization": f"Bearer {token}",
                "apikey": token,
                "Accept": "application/json",
            },
            timeout=45,
        )
        if r.status_code != 200:
            warn_skip("NSW FuelCheck", f"HTTP {r.status_code}")
            return None
        doc = r.json()
    except Exception as exc:
        warn_skip("NSW FuelCheck", f"{type(exc).__name__}: {exc}")
        return None

    prices: list[float] = []
    dates: set[str] = set()
    fuel_types = fuel_types or RETAIL_PRODUCTS["ulp91"]["nsw_fuel_types"]
    price_nodes = doc.get("prices") or doc.get("Prices") or doc.get("fuelPrices") or []
    if isinstance(price_nodes, dict):
        price_nodes = list(price_nodes.values())
    for item in price_nodes if isinstance(price_nodes, list) else []:
        if not isinstance(item, dict):
            continue
        fuel_type = str(item.get("fueltype") or item.get("fuelType") or item.get("FuelType") or "").upper()
        if fuel_type and fuel_type not in fuel_types:
            continue
        try:
            price = float(item.get("price") or item.get("Price"))
        except (TypeError, ValueError):
            continue
        if price > 0:
            prices.append(price)
        date_text = item.get("lastupdated") or item.get("lastUpdated") or item.get("TransactionDateUtc")
        if isinstance(date_text, str) and len(date_text) >= 10:
            dates.add(date_text[:10])

    source_date = sorted(dates)[-1] if dates else dt.datetime.now(dt.timezone.utc).date().isoformat()
    return retail_result("NSW", prices, source_date, f"NSW FuelCheck API {label}")


def fetch_retail_multistate(source: dict[str, Any]) -> dict[str, Any]:
    product_key = source.get("retail_product", "ulp91")
    product = RETAIL_PRODUCTS.get(product_key)
    if not product:
        raise RuntimeError(f"{source['id']}: unknown retail_product {product_key!r}")
    label = product["label"]

    today = dt.datetime.now(dt.timezone(dt.timedelta(hours=10))).date()
    qld_package_id = source.get("qld_package_id", f"fuel-price-reporting-{today.year}")
    wa_url = source.get("wa_fetch_url")
    if wa_url is None and product.get("wa_product_code") is not None:
        wa_url = f"https://www.fuelwatch.wa.gov.au/fuelwatch/fuelWatchRSS?Product={product['wa_product_code']}&Day=today"

    contributors = [
        fetch_nsw_fuelcheck(
            source.get("nsw_fetch_url", "https://api.onegov.nsw.gov.au/FuelPriceCheck/v2/fuel/prices"),
            product["nsw_fuel_types"],
            label,
        ),
        fetch_qld_open_data(
            source.get("qld_ckan_base", "https://www.data.qld.gov.au"),
            qld_package_id,
            product["qld_fuel_types"],
            label,
        ),
        fetch_wa_fuelwatch(wa_url, label) if wa_url else None,
    ]
    states = [state for state in contributors if state]
    if not states:
        return {
            "status": "unavailable",
            "unit": "cents per litre",
            "values": [],
            "last_data_point": None,
            "notes": f"No public state retail fuel feed returned usable {label} observations during this fetch.",
            "extra": {
                "schema": "retail_fuel_multistate.v1",
                "fields": {"states": [], "skipped": ["NSW", "QLD", "WA"]},
            },
        }

    total_stations = sum(state["stations"] for state in states)
    weighted = sum(state["average"] * state["stations"] for state in states) / total_stations
    latest_date = max(state["date"] for state in states if state.get("date"))
    labels = [
        f"{state['state']}: {state['stations']} stations, avg {state['average']} c/L, source date {state['date']}"
        for state in states
    ]
    return {
        "unit": "cents per litre",
        "values": [{"t": latest_date, "v": round(weighted, 1)}],
        "last_data_point": latest_date,
        "notes": (
            f"Multi-state {label} average weighted by station count. "
            + "; ".join(labels)
            + ". NSW contributes only when NSW_FUELCHECK_API_KEY is configured."
        ),
        "extra": {
            "schema": "retail_fuel_multistate.v1",
            "fields": {"product": product_key, "label": label, "states": states},
        },
    }


STATE_NAME_TO_CODE = {
    "Western Australia": "WA",
    "Queensland": "QLD",
    "Victoria": "VIC",
    "South Australia": "SA",
    "Northern Territory": "NT",
    "New South Wales": "NSW",
    "Tasmania": "TAS",
    "ACT": "ACT",
}

STATE_CODE_TO_NAME = {
    "WA": "Western Australia",
    "QLD": "Queensland",
    "VIC": "Victoria",
    "SA": "South Australia",
    "NT": "Northern Territory",
    "NSW": "New South Wales",
    "TAS": "Tasmania",
    "ACT": "Australian Capital Territory",
}


def arcgis_group_count(query_url: str, group_fields: list[str], where: str = "1=1") -> list[dict[str, Any]]:
    params = {
        "f": "json",
        "where": where,
        "returnGeometry": "false",
        "groupByFieldsForStatistics": ",".join(group_fields),
        "outStatistics": json.dumps(
            [
                {
                    "statisticType": "count",
                    "onStatisticField": "OBJECTID",
                    "outStatisticFieldName": "record_count",
                }
            ]
        ),
    }
    r = requests.get(query_url, params=params, headers={"User-Agent": UA}, timeout=45)
    r.raise_for_status()
    doc = r.json()
    if "error" in doc:
        raise RuntimeError(f"ArcGIS query failed for {query_url}: {doc['error']}")
    features = doc.get("features")
    if not isinstance(features, list):
        raise RuntimeError(f"ArcGIS query for {query_url} did not return features")
    return [feature.get("attributes", {}) for feature in features if isinstance(feature, dict)]


def arcgis_features(query_url: str, where: str, out_fields: list[str]) -> list[dict[str, Any]]:
    features: list[dict[str, Any]] = []
    offset = 0
    page_size = 2000
    while True:
        params = {
            "f": "json",
            "where": where,
            "returnGeometry": "false",
            "outFields": ",".join(out_fields),
            "resultRecordCount": page_size,
            "resultOffset": offset,
            "orderByFields": out_fields[0],
        }
        r = requests.get(query_url, params=params, headers={"User-Agent": UA}, timeout=45)
        r.raise_for_status()
        doc = r.json()
        if "error" in doc:
            raise RuntimeError(f"ArcGIS query failed for {query_url}: {doc['error']}")
        page = doc.get("features")
        if not isinstance(page, list):
            raise RuntimeError(f"ArcGIS query for {query_url} did not return features")
        features.extend(feature.get("attributes", {}) for feature in page if isinstance(feature, dict))
        if not doc.get("exceededTransferLimit") or len(page) < page_size:
            break
        offset += len(page)
    return features


def arcgis_date(value: Any) -> str | None:
    if isinstance(value, (int, float)):
        return (dt.datetime(1970, 1, 1, tzinfo=dt.timezone.utc) + dt.timedelta(milliseconds=value)).date().isoformat()
    return None


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"n/a", "na", "none", "null"}:
        return None
    return text


def clean_number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text or text.lower() in {"n/a", "na", "other"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def product_class(resource: str | None) -> str:
    text = (resource or "").lower()
    classes = []
    if "lng" in text:
        classes.append("LNG")
    if "gas" in text:
        classes.append("gas")
    if "oil" in text:
        classes.append("crude/oil")
    if "condensate" in text:
        classes.append("condensate")
    return " / ".join(dict.fromkeys(classes)) or (resource or "Unavailable")


def fetch_nopta_petroleum_counts(source: dict[str, Any]) -> dict[str, Any]:
    title_url = source.get("fetch_url")
    wells_url = source.get("nopta_wells_query_url")
    if not title_url or not wells_url:
        raise RuntimeError(f"{source['id']}: fetch_url and nopta_wells_query_url are required")

    title_rows = arcgis_group_count(
        title_url,
        ["OffShoreAr", "TitleType", "Status"],
        "TitleType <> 'Greenhouse Gas Assessment Permit'",
    )
    well_rows = arcgis_group_count(wells_url, ["OffshoreArea", "Type", "IsOffshore"])

    states = {
        code: {
            "state_code": code,
            "state_name": STATE_CODE_TO_NAME[code],
            "title_records": {
                "total_current_non_ghg": 0,
                "active": 0,
                "pending_application": 0,
                "by_title_type": {},
            },
            "well_records": {
                "total_layer_records": 0,
                "known_petroleum_type_records": 0,
                "offshore_records": 0,
                "onshore_or_unspecified_records": 0,
            },
        }
        for code in ["WA", "QLD", "VIC", "SA", "NT", "NSW", "TAS", "ACT"]
    }
    external_or_unallocated = {
        "title_records": {},
        "well_records": {},
    }

    for row in title_rows:
        area = row.get("OffShoreAr")
        count = int(row.get("record_count") or 0)
        title_type = row.get("TitleType") or "Unspecified"
        status = row.get("Status") or "Unspecified"
        code = STATE_NAME_TO_CODE.get(area)
        if code:
            target = states[code]["title_records"]
            target["total_current_non_ghg"] += count
            target["by_title_type"][title_type] = target["by_title_type"].get(title_type, 0) + count
            if status == "Active":
                target["active"] += count
            elif status == "Pending Application":
                target["pending_application"] += count
        else:
            key = area or "Unallocated"
            external_or_unallocated["title_records"][key] = (
                external_or_unallocated["title_records"].get(key, 0) + count
            )

    for row in well_rows:
        area = row.get("OffshoreArea")
        count = int(row.get("record_count") or 0)
        well_type = row.get("Type")
        is_offshore = row.get("IsOffshore")
        code = STATE_NAME_TO_CODE.get(area)
        if code:
            target = states[code]["well_records"]
            target["total_layer_records"] += count
            if well_type == "Petroleum":
                target["known_petroleum_type_records"] += count
            if is_offshore == "Yes":
                target["offshore_records"] += count
            else:
                target["onshore_or_unspecified_records"] += count
        else:
            key = area or "Unallocated"
            external_or_unallocated["well_records"][key] = (
                external_or_unallocated["well_records"].get(key, 0) + count
            )

    state_rows = []
    for code, row in states.items():
        row["title_records"]["by_title_type"] = [
            {"title_type": title_type, "count": count}
            for title_type, count in sorted(row["title_records"]["by_title_type"].items())
        ]
        row["labels"] = {
            "title_records": "Observed current NOPTA non-GHG offshore title/permit records. Active and pending rows are kept separate.",
            "well_records": "Observed NOPTA Petroleum Wells layer records by OffshoreArea. This is not an active or producing well count.",
        }
        state_rows.append(row)

    today = dt.datetime.now(dt.timezone(dt.timedelta(hours=10))).date().isoformat()
    return {
        "unit": "structured counts",
        "values": [],
        "last_data_point": today,
        "notes": (
            "Fetched from NOPTA ArcGIS REST FeatureServer public endpoints. Counts keep "
            "titles/permits, title types, status and petroleum-wells layer records separate. "
            "Greenhouse Gas Assessment Permits are excluded from petroleum title counts. "
            "Well records are not presented as active or producing wells."
        ),
        "extra": {
            "schema": "state_petroleum_nopta_counts.v1",
            "fields": {
                "as_at": today,
                "state_rows": state_rows,
                "external_or_unallocated": external_or_unallocated,
                "title_source_url": title_url,
                "wells_source_url": wells_url,
                "object_definitions": {
                    "title_records": "Current NOPTA Titles and Permits FeatureServer records, excluding Greenhouse Gas Assessment Permits.",
                    "active_title_records": "Rows where Status is Active.",
                    "pending_application_records": "Rows where Status is Pending Application.",
                    "well_records": "Rows in the NOPTA Petroleum Wells FeatureServer grouped by OffshoreArea.",
                    "known_petroleum_type_records": "Well-layer rows where Type equals Petroleum.",
                },
            },
        },
    }


def fetch_nopta_production_licence_map(source: dict[str, Any]) -> dict[str, Any]:
    title_url = source.get("fetch_url")
    if not title_url:
        raise RuntimeError(f"{source['id']}: fetch_url is required")

    rows = arcgis_features(
        title_url,
        "TitleType = 'Production Licence' AND Status = 'Active'",
        [
            "Title",
            "RelTitle",
            "TitleType",
            "ExpiryDate",
            "GrantDate",
            "Status",
            "FieldName",
            "BasinName",
            "SubBasin",
            "OffShoreAr",
            "TitleOprat",
            "TitleHold",
            "NoOfBlocks",
            "AreaKM2",
            "NEATS_Links",
        ],
    )

    state_rows = {
        code: {
            "state_code": code,
            "state_name": STATE_CODE_TO_NAME[code],
            "production_licence_records": 0,
            "mapped_field_records": 0,
            "mapped_operator_records": 0,
            "basins": set(),
            "operators": set(),
        }
        for code in ["WA", "QLD", "VIC", "SA", "NT", "NSW", "TAS", "ACT"]
    }
    mapped_rows: list[dict[str, Any]] = []
    external_or_unallocated: list[dict[str, Any]] = []

    for row in rows:
        offshore_area = clean_text(row.get("OffShoreAr"))
        state_code = STATE_NAME_TO_CODE.get(offshore_area)
        title_holders_raw = clean_text(row.get("TitleHold"))
        item = {
            "state_code": state_code,
            "state_name": STATE_CODE_TO_NAME.get(state_code) if state_code else offshore_area,
            "offshore_area": offshore_area,
            "basin_name": clean_text(row.get("BasinName")),
            "sub_basin": clean_text(row.get("SubBasin")),
            "title": clean_text(row.get("Title")),
            "related_title": clean_text(row.get("RelTitle")),
            "field_name": clean_text(row.get("FieldName")),
            "title_operator": clean_text(row.get("TitleOprat")),
            "title_holders_raw": title_holders_raw,
            "title_holders": [part.strip() for part in title_holders_raw.split(",") if part.strip()] if title_holders_raw else [],
            "title_type": clean_text(row.get("TitleType")),
            "status": clean_text(row.get("Status")),
            "grant_date": arcgis_date(row.get("GrantDate")),
            "expiry_date": arcgis_date(row.get("ExpiryDate")),
            "blocks": int(row["NoOfBlocks"]) if clean_number(row.get("NoOfBlocks")) is not None else None,
            "area_km2": round(clean_number(row.get("AreaKM2")), 2) if clean_number(row.get("AreaKM2")) is not None else None,
            "neats_url": clean_text(row.get("NEATS_Links")),
            "product_class": "petroleum",
            "production_metric_name": "Active offshore petroleum production licence record",
            "production_metric_value": None,
            "production_unit": None,
            "production_period": "current NOPTA title record",
            "trust_label": "Observed",
            "mapping_status": "Mapped active production licence; source does not publish production volume in this layer.",
            "notes": "NOPTA title/operator/field metadata. This is not a production-volume, revenue or project-output measure.",
        }
        if state_code:
            target = state_rows[state_code]
            target["production_licence_records"] += 1
            if item["field_name"]:
                target["mapped_field_records"] += 1
            if item["title_operator"]:
                target["mapped_operator_records"] += 1
                target["operators"].add(item["title_operator"])
            if item["basin_name"]:
                target["basins"].add(item["basin_name"])
            mapped_rows.append(item)
        else:
            external_or_unallocated.append(item)

    summary_rows = []
    for row in state_rows.values():
        summary = dict(row)
        summary["basins"] = sorted(summary["basins"])
        summary["operators"] = sorted(summary["operators"])
        summary_rows.append(summary)

    today = dt.datetime.now(dt.timezone(dt.timedelta(hours=10))).date().isoformat()
    return {
        "unit": "active offshore petroleum production licence records",
        "values": [{"t": today, "v": len(mapped_rows)}],
        "last_data_point": today,
        "notes": (
            "Fetched from the NOPTA Titles and Permits Current FeatureServer for active "
            "petroleum production licence records. Rows map offshore area/state, basin, "
            "field name, title, title operator and title holders where published. This "
            "is partial offshore regulatory mapping and does not contain production volumes."
        ),
        "extra": {
            "schema": "state_petroleum_production_licence_map.v1",
            "fields": {
                "as_at": today,
                "source_scope": "Active NOPTA Production Licence rows only.",
                "state_rows": summary_rows,
                "production_licence_rows": mapped_rows,
                "external_or_unallocated": external_or_unallocated,
                "title_source_url": title_url,
                "mapping_definitions": {
                    "state": "Derived from NOPTA OffShoreAr where it matches an Australian state or territory.",
                    "basin": "NOPTA BasinName field.",
                    "field_name": "NOPTA FieldName field. It can contain one or multiple fields and is not split or inferred.",
                    "operator": "NOPTA TitleOprat field.",
                    "title_holders": "NOPTA TitleHold field, preserved as raw text and a best-effort split list.",
                    "production_metric_value": "Unavailable. The source layer publishes regulatory title metadata, not production volume.",
                },
            },
        },
    }


def find_workbook_header(rows: list[tuple[Any, ...]], required: list[str]) -> tuple[int, list[str]]:
    normalized_required = {item.lower() for item in required}
    for idx, row in enumerate(rows):
        header = [str(cell).strip() if cell is not None else "" for cell in row]
        normalized = {cell.lower() for cell in header if cell}
        if normalized_required.issubset(normalized):
            return idx, header
    raise RuntimeError(f"Workbook sheet did not contain required columns: {', '.join(required)}")


def fetch_remp_oil_gas_projects(source: dict[str, Any]) -> dict[str, Any]:
    url = source.get("fetch_url")
    if not url:
        raise RuntimeError(f"{source['id']}: fetch_url is required")
    user_agent = source.get("fetch_user_agent")
    if user_agent == "browser-compatible":
        user_agent = "Mozilla/5.0"
    wb = load_xlsx_from_url(
        url,
        {
            "User-Agent": user_agent or "Mozilla/5.0",
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
        },
    )
    sheet_name = source.get("remp_sheet", "Oil & gas")
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"REMP workbook did not contain sheet {sheet_name!r}")
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    header_idx, header = find_workbook_header(rows, ["Project", "Company", "State", "Resource"])
    header_map = {name: idx for idx, name in enumerate(header) if name}

    project_rows: list[dict[str, Any]] = []
    state_rows = {
        code: {
            "state_code": code,
            "state_name": STATE_CODE_TO_NAME[code],
            "project_rows": 0,
            "committed_rows": 0,
            "publicly_announced_rows": 0,
            "mapped_company_rows": 0,
            "product_classes": set(),
        }
        for code in ["WA", "QLD", "VIC", "SA", "NT", "NSW", "TAS", "ACT"]
    }

    def cell(row: tuple[Any, ...], name: str) -> Any:
        idx = header_map.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    for raw in rows[header_idx + 1:]:
        project_name = clean_text(cell(raw, "Project"))
        if not project_name:
            continue
        state_code = clean_text(cell(raw, "State"))
        if state_code not in state_rows:
            continue
        resource = clean_text(cell(raw, "Resource"))
        status = clean_text(cell(raw, "Status"))
        capacity_value = clean_number(cell(raw, "Annual Estimated New Capacity"))
        capacity_unit = clean_text(cell(raw, "Capacity Unit"))
        cost_value = clean_number(cell(raw, "Cost Estimate A$m"))
        row = {
            "state_code": state_code,
            "state_name": STATE_CODE_TO_NAME[state_code],
            "basin_name": None,
            "project_name": project_name,
            "field_name": None,
            "operator_name": None,
            "company_name": clean_text(cell(raw, "Company")),
            "product_class": product_class(resource),
            "resource": resource,
            "project_type": clean_text(cell(raw, "Type")),
            "status": status,
            "latitude": clean_number(cell(raw, "Latitude")),
            "longitude": clean_number(cell(raw, "Longitude")),
            "production_metric_name": "Annual estimated new capacity (not current production)",
            "production_metric_value": capacity_value,
            "production_unit": capacity_unit,
            "production_period": "Resources and Energy Major Projects 2024",
            "cost_estimate_aud_m": cost_value,
            "estimated_start_commercial_operation": clean_text(cell(raw, "Estimated Start Commercial Operation")),
            "trust_label": "Partial coverage",
            "mapping_status": "Major project/development row; not a current production-volume row.",
            "notes": "REMP maps project, company, state, resource and estimated new capacity where published. It does not publish basin, operator role or current production volume.",
        }
        project_rows.append(row)
        summary = state_rows[state_code]
        summary["project_rows"] += 1
        if status == "Committed":
            summary["committed_rows"] += 1
        elif status == "Publicly_announced":
            summary["publicly_announced_rows"] += 1
        if row["company_name"]:
            summary["mapped_company_rows"] += 1
        summary["product_classes"].add(row["product_class"])

    if not project_rows:
        raise RuntimeError("REMP Oil & gas sheet contained no project rows")

    summary_rows = []
    for row in state_rows.values():
        summary = dict(row)
        summary["product_classes"] = sorted(summary["product_classes"])
        summary_rows.append(summary)

    return {
        "unit": "oil and gas major project rows",
        "values": [{"t": "2024-12-20", "v": len(project_rows)}],
        "last_data_point": "2024-12-20",
        "notes": (
            "Parsed the Department of Industry, Science and Resources Resources and Energy "
            "Major Projects 2024 data workbook, Oil & gas sheet. Rows map project, company, "
            "state, resource, status and annual estimated new capacity where published. "
            "This is a major-project/development list, not current production by company."
        ),
        "source_url_resolved": url,
        "extra": {
            "schema": "state_oil_gas_major_projects_remp.v1",
            "fields": {
                "report": "Resources and Energy Major Projects 2024",
                "publication_date": "2024-12-20",
                "source_period_note": "Publication page states the 2024 report updates project developments from November 2023 to October 2024; the workbook table-of-contents text still refers to project data as at 31 October 2023.",
                "sheet": sheet_name,
                "state_rows": summary_rows,
                "project_rows": project_rows,
                "mapping_definitions": {
                    "project": "REMP Project column.",
                    "company": "REMP Company column. The row does not distinguish operator, owner, proponent or joint-venture role.",
                    "state": "REMP State column.",
                    "product_class": "Derived only from the REMP Resource text for display grouping.",
                    "production_metric_value": "Annual Estimated New Capacity where numeric. It is not current production.",
                    "basin_name": "Unavailable; the REMP Oil & gas sheet does not publish a basin column.",
                },
            },
        },
    }


Fetcher = Callable[[dict[str, Any]], dict[str, Any]]
FETCHERS: dict[str, Fetcher] = {
    "eia_brent": lambda source: fetch_eia_series("RBRTE", source.get("fetch_url")),
    "eia_wti": lambda source: fetch_eia_series("RWTC", source.get("fetch_url")),
    "eia_diesel_retail_us": fetch_fred_fuel_csv,
    "eia_jet_spot_usgc": fetch_fred_fuel_csv,
    "aps_monthly": fetch_aps_xlsx_series,
    "aps_production_petrol": fetch_aps_xlsx_series,
    "aps_production_diesel": fetch_aps_xlsx_series,
    "aps_production_jet": fetch_aps_xlsx_series,
    "aps_production_fuel_oil": fetch_aps_xlsx_series,
    "abs_petroleum_imports": lambda source: fetch_abs_sdmx(
        source["fetch_dataflow"],
        source["fetch_key"],
        source.get("fetch_start_period"),
        source.get("sdmx_note_suffix"),
        source.get("fetch_url"),
    ),
    "abs_petroleum_imports_yoy": fetch_abs_petroleum_imports_yoy,
    "abs_fertiliser_imports": lambda source: fetch_abs_sdmx(
        source["fetch_dataflow"],
        source["fetch_key"],
        source.get("fetch_start_period"),
        source.get("sdmx_note_suffix"),
        source.get("fetch_url"),
    ),
    "abs_unemployment_rate": lambda source: fetch_abs_sdmx(
        source["fetch_dataflow"],
        source["fetch_key"],
        source.get("fetch_start_period"),
        source.get("sdmx_note_suffix"),
        source.get("fetch_url"),
    ),
    "abs_cpi_inflation": lambda source: fetch_abs_sdmx(
        source["fetch_dataflow"],
        source["fetch_key"],
        source.get("fetch_start_period"),
        source.get("sdmx_note_suffix"),
        source.get("fetch_url"),
    ),
    "abs_gdp_real_growth": lambda source: fetch_abs_sdmx(
        source["fetch_dataflow"],
        source["fetch_key"],
        source.get("fetch_start_period"),
        source.get("sdmx_note_suffix"),
        source.get("fetch_url"),
    ),
    "abs_population_quarterly": lambda source: fetch_abs_sdmx(
        source["fetch_dataflow"],
        source["fetch_key"],
        source.get("fetch_start_period"),
        source.get("sdmx_note_suffix"),
        source.get("fetch_url"),
    ),
    "abs_population_growth_rate": lambda source: fetch_abs_sdmx(
        source["fetch_dataflow"],
        source["fetch_key"],
        source.get("fetch_start_period"),
        source.get("sdmx_note_suffix"),
        source.get("fetch_url"),
    ),
    "abs_residential_dwelling_stock": lambda source: fetch_abs_sdmx(
        source["fetch_dataflow"],
        source["fetch_key"],
        source.get("fetch_start_period"),
        source.get("sdmx_note_suffix"),
        source.get("fetch_url"),
    ),
    "abs_manufacturing_employment": fetch_abs_release_xlsx_series,
    "abs_manufacturing_output_index": fetch_abs_release_xlsx_series,
    "abs_manufactured_exports_total": fetch_abs_release_xlsx_series,
    "abs_manufacturing_capex": fetch_abs_release_xlsx_series,
    "abs_food_beverage_employment": fetch_abs_release_xlsx_series,
    "abs_fertiliser_source_concentration": fetch_abs_fertiliser_source_concentration,
    "rba_aud_usd": lambda source: fetch_rba_f11(source["fetch_url"]),
    "rba_cash_rate": lambda source: fetch_rba_csv_column(
        source["fetch_url"],
        title_substring="Cash Rate Target",
        aggregate="monthly_mean",
        unit="per cent",
        notes="Monthly mean of daily RBA cash-rate target observations from Statistical Table F1.1. Trimmed to last 60 months.",
        ndigits=2,
    ),
    "rba_household_debt_to_income": lambda source: fetch_rba_csv_column(
        source["fetch_url"],
        title_substring="Household debt to income",
        aggregate="quarter_end",
        unit="per cent",
        notes="Quarter-end value of household debt as a per cent of annualised household disposable income, from RBA Statistical Table E2. Trimmed to last 40 quarters.",
        ndigits=1,
    ),
    "aemo_nem_average_wholesale_price": lambda source: fetch_aemo_nem_price_demand("price"),
    "aemo_nem_total_demand": lambda source: fetch_aemo_nem_price_demand("demand"),
    "aus_retail_fuel_multistate": fetch_retail_multistate,
    "aip_tgp": fetch_aip_tgp,
    "qld_fuel_security_unavailable_reports": fetch_qld_unavailable_reports,
    "state_petroleum_nopta_counts": fetch_nopta_petroleum_counts,
    "state_petroleum_production_licence_map": fetch_nopta_production_licence_map,
    "state_oil_gas_major_projects_remp": fetch_remp_oil_gas_projects,
}


def load_sources() -> list[dict[str, Any]]:
    with SOURCES_FILE.open("r", encoding="utf-8") as f:
        doc = yaml.safe_load(f)
    return doc.get("sources", [])


def write_generated(source: dict[str, Any], block: dict[str, Any]) -> pathlib.Path:
    now = dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")
    status = block.get("status", "ok")
    payload = {
        "series_id": source["id"],
        "source_id": source["id"],
        "source_name": source["human_name"],
        "source_url": source.get("canonical_url") or source["url"],
        "unit": block["unit"],
        "retrieved_at": now if status == "ok" else None,
        "last_data_point": block.get("last_data_point"),
        "values": block["values"],
        "notes": block.get("notes", ""),
        "status": status,
        "manual_entry": False,
    }
    if source.get("rights"):
        payload["source_rights"] = source["rights"]
    if source.get("citation"):
        payload["citation"] = source["citation"]
    if block.get("extra") is not None:
        payload["extra"] = block["extra"]

    path = GENERATED_DIR / f"{source['id']}.json"
    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
        f.write("\n")
    return path


def check_url(url: str) -> tuple[bool, str]:
    try:
        r = requests.head(url, headers={"User-Agent": UA}, timeout=20, allow_redirects=True)
        if r.status_code >= 400:
            r = requests.get(url, headers={"User-Agent": UA}, timeout=20, allow_redirects=True, stream=True)
            r.close()
        if r.status_code >= 400:
            return False, f"HTTP {r.status_code}"
        return True, f"HTTP {r.status_code}"
    except requests.RequestException as e:
        return False, f"{type(e).__name__}: {e}"


def source_check_url(source: dict[str, Any], *, all_links: bool) -> str:
    if all_links:
        return source.get("canonical_url") or source["url"]
    return source.get("fetch_url") or source.get("canonical_url") or source["url"]


def main() -> int:
    ap = argparse.ArgumentParser(description="Fetch data for Fuel Resilience AU")
    ap.add_argument("--only", help="Only fetch this source id")
    ap.add_argument("--check", action="store_true", help="Check programmatic fetch URLs only")
    ap.add_argument("--check-all-links", action="store_true", help="Check every canonical source URL")
    args = ap.parse_args()

    sources = load_sources()
    if not sources:
        print("No sources found in data/sources.yml", file=sys.stderr)
        return 2

    errors: list[str] = []
    wrote: list[str] = []
    skipped: list[tuple[str, str]] = []

    derived_by_parent: dict[str, list[dict[str, Any]]] = {}
    for source in sources:
        if source.get("fetch") == "derived" and source.get("derived_from"):
            derived_by_parent.setdefault(source["derived_from"], []).append(source)

    wrote_ids: set[str] = set()

    def fetch_and_write(source: dict[str, Any]) -> None:
        sid = source["id"]
        if sid in wrote_ids:
            return
        fetcher = FETCHERS.get(sid)
        if fetcher is None and source.get("aps_sheet") and source.get("aps_column"):
            fetcher = fetch_aps_xlsx_series
        if fetcher is None and source.get("retail_product"):
            fetcher = fetch_retail_multistate
        if fetcher is None and source.get("derived_from") and source.get("derived_field"):
            fetcher = fetch_extra_field_derived
        if not fetcher:
            errors.append(f"{sid}: marked {source.get('fetch')} but no fetcher registered")
            return
        try:
            block = fetcher(source)
            path = write_generated(source, block)
            wrote_ids.add(sid)
            wrote.append(str(path.relative_to(ROOT)))
            print(f"[OK ] {sid:<32} -> {path.relative_to(ROOT)}")
        except Exception as e:
            errors.append(f"{sid}: {type(e).__name__}: {e}")
            print(f"[ERR] {sid:<32} {e}", file=sys.stderr)

    for source in sources:
        sid = source["id"]
        if args.only and sid != args.only:
            continue
        fetch_mode = source.get("fetch", "manual")

        if args.check or args.check_all_links:
            if args.check and fetch_mode != "programmatic":
                skipped.append((sid, "not programmatic; skipped blocking fetch check"))
                continue
            url = source_check_url(source, all_links=args.check_all_links)
            ok, msg = check_url(url)
            tag = "OK " if ok else "ERR"
            print(f"[{tag}] {sid:<32} {msg}  {url}")
            if not ok:
                errors.append(f"{sid}: {msg}")
            continue

        if fetch_mode in {"programmatic", "derived"}:
            fetch_and_write(source)
            if fetch_mode == "programmatic":
                for derived in derived_by_parent.get(sid, []):
                    fetch_and_write(derived)
        elif fetch_mode == "manual":
            manual_path = MANUAL_DIR / f"{sid}.json"
            skipped.append((sid, "manual, file present" if manual_path.exists() else "manual, FILE MISSING"))
        elif fetch_mode == "unavailable":
            skipped.append((sid, "unavailable by design"))
        else:
            errors.append(f"{sid}: unknown fetch mode '{fetch_mode}'")

    print()
    print(f"Wrote   {len(wrote)} generated file(s)")
    print(f"Skipped {len(skipped)} source(s)")
    for sid, why in skipped:
        print(f"        - {sid}: {why}")

    if errors:
        print()
        print(f"ERRORS ({len(errors)}):", file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
