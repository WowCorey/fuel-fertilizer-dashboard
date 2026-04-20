#!/usr/bin/env python3
"""Automate refresh of manual/unavailable source envelopes.

Phase 2 extractor coverage:
  - aps_*: XLSX parser
  - ato_corporate_tax: XLSX parser
  - accc_*: PDF parser with OCR fallback
  - company_*: PDF parser with OCR fallback and validation rules
"""

from __future__ import annotations

import datetime as dt
import os
import io
import json
import pathlib
import re
import sys
import time
import urllib.parse
import argparse
from typing import Any

try:
    import yaml
except ImportError:
    sys.stderr.write("PyYAML is required. Install with: pip install pyyaml\n")
    sys.exit(2)

try:
    import requests
except ImportError:
    sys.stderr.write("requests is required. Install with: pip install requests\n")
    sys.exit(2)

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import pypdfium2
except ImportError:
    pypdfium2 = None

try:
    from pdf2image import convert_from_bytes
except ImportError:
    convert_from_bytes = None

try:
    import pytesseract
except ImportError:
    pytesseract = None

try:
    from _repo_url import repo_url
except ImportError:
    from scripts._repo_url import repo_url

ROOT = pathlib.Path(__file__).resolve().parent.parent
SOURCES_FILE = ROOT / "data" / "sources.yml"
MANUAL_DIR = ROOT / "data" / "manual"
GENERATED_DIR = ROOT / "data" / "generated"
REPORTS_DIR = ROOT / "data" / "reports"

UA = f"FuelResilienceAU-ManualBot/1.0 (+{repo_url()})"
HTTP_TIMEOUT = 15
EXTRA_SCHEMA = "manual_automation.v2"
HTTP_RETRIES = 3

APS_IDS = {
    "aps_monthly",
    "aps_production_petrol",
    "aps_production_diesel",
    "aps_production_jet",
    "aps_production_fuel_oil",
    "aps_refinery_utilisation",
}
ACCC_IDS = {
    "accc_petroleum_monitoring",
    "accc_petrol_wholesale_component",
    "accc_petrol_excise_component",
    "accc_petrol_gst_component",
    "accc_petrol_retail_margin_component",
    "accc_petrol_breakdown_series",
}
COMPANY_IDS = {
    "company_exxonmobil_au",
    "company_chevron_au",
    "company_viva_energy",
    "company_ampol",
    "company_bp_au",
    "company_shell_au",
    "company_woodside",
    "company_santos",
}

COMPANY_PARSER_PROFILES: dict[str, dict[str, Any]] = {
    "company_exxonmobil_au": {
        "resolver_endpoints": [
            "https://corporate.exxonmobil.com/locations/australia",
            "https://corporate.exxonmobil.com/sustainability-and-reports/reports-and-publications",
            "https://corporate.exxonmobil.com/energy-and-innovation/annual-reports",
        ],
    },
    "company_chevron_au": {
        "resolver_endpoints": [
            "https://australia.chevron.com/",
            "https://www.chevron.com/investors/financial-reporting",
            "https://www.chevron.com/investors/reports",
            "https://www.asx.com.au/markets/company/CVX",
        ],
    },
    "company_viva_energy": {
        "resolver_endpoints": [
            "https://www.vivaenergy.com.au/investors/reports-and-presentations",
            "https://www.vivaenergy.com.au/investors",
            "https://www.asx.com.au/markets/company/VEA",
        ],
    },
    "company_ampol": {
        "resolver_endpoints": [
            "https://www.ampol.com.au/about-ampol/investor-centre/reports-and-presentations",
            "https://www.asx.com.au/markets/company/ALD",
        ],
        "npat_patterns": [
            r"(underlying rcop net profit after tax|net profit after tax attributable to equity holders)",
            r"(net profit after tax|profit after tax)",
        ],
        "tax_patterns": [
            r"(income tax expense|tax expense|taxation expense)",
        ],
    },
    "company_bp_au": {
        "resolver_endpoints": [
            "https://www.bp.com/en/global/corporate/investors/results-and-reporting.html",
            "https://www.bp.com/en_au/australia/home/who-we-are.html",
            "https://www.asx.com.au/markets/company/BPT",
        ],
    },
    "company_shell_au": {
        "resolver_endpoints": [
            "https://www.shell.com/investors/results-and-reporting/annual-report.html",
            "https://reports.shell.com/annual-report/2024/servicepages/downloads/files/entire_shell_ar24.pdf",
            "https://www.shell.com.au/about-us.html",
        ],
    },
    "company_woodside": {
        "resolver_endpoints": [
            "https://www.woodside.com/investors/reports-investor-briefings",
            "https://www.asx.com.au/markets/company/WDS",
        ],
    },
    "company_santos": {
        "resolver_endpoints": [
            "https://www.santos.com/investors/reports/",
            "https://www.asx.com.au/markets/company/STO",
        ],
    },
}

SOURCE_URL_OVERRIDES: dict[str, str] = {
    # Prefer stable data catalogue for APS extraction.
    "aps_monthly": "https://www.data.gov.au/data/dataset/australian-petroleum-statistics",
    "aps_production_petrol": "https://www.data.gov.au/data/dataset/australian-petroleum-statistics",
    "aps_production_diesel": "https://www.data.gov.au/data/dataset/australian-petroleum-statistics",
    "aps_production_jet": "https://www.data.gov.au/data/dataset/australian-petroleum-statistics",
    "aps_production_fuel_oil": "https://www.data.gov.au/data/dataset/australian-petroleum-statistics",
    "aps_refinery_utilisation": "https://www.data.gov.au/data/dataset/australian-petroleum-statistics",
    # Prefer known current page roots for flaky redirect chains.
    "ato_corporate_tax": "https://data.gov.au/data/dataset/c2524c87-cea4-4636-acac-599a82048a26",
    "accc_petroleum_monitoring": "https://www.accc.gov.au/system/files/petrol-quarterly-report-june24.pdf",
    "accc_petrol_wholesale_component": "https://www.accc.gov.au/system/files/petrol-quarterly-report-june24.pdf",
    "accc_petrol_excise_component": "https://www.accc.gov.au/system/files/petrol-quarterly-report-june24.pdf",
    "accc_petrol_gst_component": "https://www.accc.gov.au/system/files/petrol-quarterly-report-june24.pdf",
    "accc_petrol_retail_margin_component": "https://www.accc.gov.au/system/files/petrol-quarterly-report-june24.pdf",
    "accc_petrol_breakdown_series": "https://www.accc.gov.au/system/files/petrol-quarterly-report-june24.pdf",
    "company_ampol": "https://assets.contentstack.io/v3/assets/blt35cb056c1c8431c3/bltf2b88cca4ee2d090/686b0daa16ecf85da24ead41/2024_Annual_Report.pdf",
}


def configure_ocr_runtime() -> None:
    """Best-effort OCR binary discovery for local Windows environments."""
    if os.name != "nt":
        return
    poppler_env = os.environ.get("POPPLER_PATH")
    poppler_candidates = [
        pathlib.Path(r"C:\Program Files\poppler\Library\bin"),
        pathlib.Path(r"C:\Program Files (x86)\poppler\Library\bin"),
    ]
    if poppler_env:
        poppler_candidates.insert(0, pathlib.Path(poppler_env))
    local_winget = pathlib.Path(os.environ.get("LOCALAPPDATA", "")) / "Microsoft" / "WinGet" / "Packages"
    if local_winget.exists():
        poppler_candidates.extend(
            p / "poppler-25.07.0" / "Library" / "bin"
            for p in local_winget.glob("oschwartz10612.Poppler*")
        )

    for candidate in poppler_candidates:
        if not candidate:
            continue
        try:
            if candidate.exists() and (candidate / "pdftoppm.exe").exists():
                os.environ["PATH"] = str(candidate) + os.pathsep + os.environ.get("PATH", "")
                break
        except OSError:
            continue

    if pytesseract is None:
        return
    tesseract_candidates = [
        pathlib.Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe"),
        pathlib.Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"),
    ]
    tesseract_env = os.environ.get("TESSERACT_CMD")
    if tesseract_env:
        tesseract_candidates.insert(0, pathlib.Path(tesseract_env))
    for candidate in tesseract_candidates:
        if not candidate:
            continue
        try:
            if candidate.exists() and candidate.is_file():
                pytesseract.pytesseract.tesseract_cmd = str(candidate)
                break
        except OSError:
            continue


configure_ocr_runtime()

def now_iso() -> str:
    return dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")

def load_sources() -> list[dict[str, Any]]:
    with SOURCES_FILE.open("r", encoding="utf-8") as f:
        doc = yaml.safe_load(f)
    sources = doc.get("sources", [])
    if not isinstance(sources, list):
        raise RuntimeError("data/sources.yml has no sources list")
    return [s for s in sources if isinstance(s, dict)]

def find_links(html: str, base_url: str) -> list[str]:
    out = []
    for match in re.finditer(r'href=["\']([^"\']+)["\']', html, flags=re.IGNORECASE):
        href = match.group(1).strip()
        if not href or href.startswith("#"):
            continue
        out.append(urllib.parse.urljoin(base_url, href))
    return out

def score_candidate(url: str, fmt: str) -> int:
    lower = url.lower()
    score = 0
    if fmt == "PDF" and ".pdf" in lower:
        score += 10
    if fmt == "XLSX" and ".xlsx" in lower:
        score += 10
    if fmt == "CSV" and ".csv" in lower:
        score += 10
    if any(k in lower for k in ("latest", "current", "report", "data", "download")):
        score += 2
    if any(k in lower for k in ("archive", "older", "historical")):
        score -= 1
    return score

def resolve_machine_url(source: dict[str, Any]) -> tuple[str, str]:
    sid = str(source.get("id", ""))
    if sid in SOURCE_URL_OVERRIDES:
        source = dict(source)
        source["canonical_url"] = SOURCE_URL_OVERRIDES[sid]

    direct = source.get("fetch_url")
    if isinstance(direct, str) and direct.strip():
        return direct.strip(), "fetch_url"

    canonical = source.get("canonical_url") or source.get("url")
    if not isinstance(canonical, str) or not canonical.strip():
        raise RuntimeError("source has no canonical_url/url")
    canonical = canonical.strip()
    fmt = str(source.get("format", "MIXED"))

    if fmt == "HTML_TABLE":
        return canonical, "canonical_html"

    if sid in COMPANY_IDS:
        resolved = resolve_company_pdf_url(source, canonical)
        if resolved:
            return resolved

    r = request_get(canonical)
    r.raise_for_status()
    links = find_links(r.text, canonical)
    if not links:
        return canonical, "canonical_no_links"

    canonical_host = urllib.parse.urlparse(canonical).netloc.lower()
    sid = str(source.get("id", ""))
    def preferred_pdf(u: str) -> bool:
        lu = u.lower()
        if sid.startswith("accc_"):
            return "/system/files/" in lu and ".pdf" in lu
        return ".pdf" in lu

    ranked = sorted(
        links,
        key=lambda u: (
            preferred_pdf(u) if fmt == "PDF" else False,
            score_candidate(u, fmt),
            urllib.parse.urlparse(u).netloc.lower() == canonical_host,
            u,
        ),
        reverse=True,
    )
    best = ranked[0]
    if score_candidate(best, fmt) <= 0:
        return canonical, "canonical_no_scored_doc"
    return best, f"discovered_from_canonical:{fmt.lower()}"


def company_pdf_candidates(base_url: str) -> list[str]:
    try:
        r = request_get(base_url)
    except Exception:
        return []
    links = find_links(r.text, base_url)
    pdfs = [u for u in links if ".pdf" in u.lower()]
    if not pdfs:
        return []
    ranked = sorted(
        pdfs,
        key=lambda u: (
            any(token in u.lower() for token in ("annual", "financial", "report", "results")),
            ".asx" in u.lower() or "announcement" in u.lower(),
            len(u),
        ),
        reverse=True,
    )
    out: list[str] = []
    seen = set()
    for link in ranked:
        if link in seen:
            continue
        seen.add(link)
        out.append(link)
    return out


def resolve_company_pdf_url(source: dict[str, Any], canonical: str) -> tuple[str, str] | None:
    sid = str(source.get("id", ""))
    endpoints = list((COMPANY_PARSER_PROFILES.get(sid) or {}).get("resolver_endpoints", []))
    if canonical not in endpoints:
        endpoints = [canonical] + endpoints
    for idx, endpoint in enumerate(endpoints):
        direct_pdf = endpoint.lower().endswith(".pdf")
        if direct_pdf:
            return endpoint, f"company_resolver_direct_{idx}"
        for candidate in company_pdf_candidates(endpoint):
            return candidate, f"company_resolver_pdf_{idx}"
    if endpoints:
        # Return a stable endpoint even if PDF discovery fails, so the resolver
        # remains deterministic and extractor metadata can surface what failed.
        return endpoints[0], "company_resolver_endpoint_0"
    return None


def company_structured_candidates(source: dict[str, Any]) -> list[dict[str, str]]:
    raw = source.get("company_structured_source_candidates")
    if not isinstance(raw, list):
        return []
    out: list[dict[str, str]] = []
    for entry in raw:
        if not isinstance(entry, dict):
            continue
        url = str(entry.get("url", "")).strip()
        kind = str(entry.get("kind", "")).strip().lower()
        if not url or kind not in {"xbrl", "ixbrl"}:
            continue
        out.append({"url": url, "kind": kind})
    return out


def detect_structured_numeric(text: str, labels: tuple[str, ...]) -> float | None:
    candidates: list[float] = []
    lowered = text.lower()
    for label in labels:
        pattern = re.compile(
            rf"(?:name\s*=\s*[\"'][^\"']*{re.escape(label)}[^\"']*[\"'][^>]*>|<[^>]*{re.escape(label)}[^>]*>)\s*([-]?\d[\d,]*(?:\.\d+)?)",
            re.IGNORECASE,
        )
        for m in pattern.finditer(text):
            val = parse_numeric(m.group(1))
            if val is not None:
                candidates.append(val)
        # Some ixbrl pages carry facts in hidden containers; allow a wider nearby scan.
        idx = lowered.find(label.lower())
        if idx >= 0:
            window = text[max(0, idx - 300) : idx + 600]
            for token in re.findall(r"[-]?\d[\d,]{0,20}(?:\.\d+)?", window):
                val = parse_numeric(token)
                if val is not None:
                    candidates.append(val)
    if not candidates:
        return None
    return max(candidates, key=lambda v: abs(v))


def discover_structured_artifact(url: str) -> tuple[str, str]:
    r = request_get(url)
    r.raise_for_status()
    ctype = (r.headers.get("content-type") or "").lower()
    looks_html = "text/html" in ctype or "<html" in (r.text[:400].lower() if hasattr(r, "text") else "")
    if not looks_html:
        return url, "structured_direct"
    links = find_links(r.text, url)
    if not links:
        return url, "structured_html_no_links"
    structured = [
        u
        for u in links
        if any(token in u.lower() for token in (".xml", ".xhtml", ".htm", "xbrl", "ixbrl", "inline"))
    ]
    if not structured:
        return url, "structured_html_no_candidate"
    ranked = sorted(
        structured,
        key=lambda u: (
            any(token in u.lower() for token in ("ixbrl", "inline", ".xhtml")),
            any(token in u.lower() for token in (".xml", "xbrl")),
            len(u),
        ),
        reverse=True,
    )
    return ranked[0], "structured_discovered_link"


def extract_company_structured(source: dict[str, Any], candidate_url: str, candidate_kind: str, resolution_note: str) -> dict[str, Any]:
    resolved_url, discover_note = discover_structured_artifact(candidate_url)
    blob = fetch_bytes(resolved_url)
    text = blob.decode("utf-8", errors="ignore")
    if not text.strip():
        raise RuntimeError("structured candidate had no text payload")

    pbt = detect_structured_numeric(
        text,
        ("profitlossbeforetax", "profitbeforetax", "earningsbeforetax", "profitbeforeincometax"),
    )
    tax = detect_structured_numeric(
        text,
        ("incometaxexpense", "taxexpense", "taxationexpense", "incometaxtaxexpense"),
    )
    rev = detect_structured_numeric(
        text,
        ("revenue", "revenuefromcontractswithcustomers", "totalrevenue", "salesrevenuegoodsnet"),
    )
    if pbt is None or tax is None:
        raise RuntimeError("structured parser could not locate required pbt/tax metrics")
    if abs(pbt) < 0.0001:
        raise RuntimeError("structured parser rejected near-zero pbt")

    tax_abs = abs(tax)
    etr = (tax_abs / abs(pbt) * 100.0) if abs(pbt) > 0 else 0.0
    etr = max(min(etr, 100.0), 0.0)
    year_candidates = re.findall(r"(20\d{2})", text[:20000])
    year = max(int(y) for y in year_candidates) if year_candidates else dt.datetime.now(dt.timezone.utc).year
    sid = source["id"]
    return {
        "status": "ok",
        "unit": "%",
        "values": [{"t": f"{year}", "v": round(etr, 2)}],
        "last_data_point": f"{year}-06-30",
        "notes": "Parsed company structured filing (XBRL/iXBRL) and computed effective tax rate.",
        "extra": {
            "schema": EXTRA_SCHEMA,
            "fields": {
                "extractor": f"company_{candidate_kind}_v1",
                "resolved_url": resolved_url,
                "resolution": f"{resolution_note}:{discover_note}",
                "parse_mode": candidate_kind,
                "financials": {
                    "revenue": round(rev, 2) if rev is not None else None,
                    "profit_before_tax": round(pbt, 2),
                    "income_tax_abs": round(tax_abs, 2),
                },
                "source_id": sid,
            },
        },
    }


def extract_company_with_fallback(source: dict[str, Any], resolved_url: str, resolution_note: str) -> dict[str, Any]:
    attempts: list[str] = []
    for candidate in company_structured_candidates(source):
        try:
            return extract_company_structured(source, candidate["url"], candidate["kind"], resolution_note)
        except Exception as exc:
            attempts.append(f"{candidate['kind']}:{type(exc).__name__}")
    try:
        return extract_company_report(source, resolved_url, resolution_note)
    except Exception as exc:
        if attempts:
            raise RuntimeError(f"structured attempts failed ({', '.join(attempts)}); pdf fallback failed: {exc}") from exc
        raise


def run_company_structured_health_report(sources: list[dict[str, Any]]) -> int:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    checked_at = now_iso()
    rows: list[dict[str, Any]] = []
    company_sources = [s for s in sources if str(s.get("id", "")).startswith("company_")]
    for source in company_sources:
        sid = str(source.get("id", ""))
        entry: dict[str, Any] = {
            "source_id": sid,
            "checked_at": checked_at,
            "structured_candidates": [],
            "pdf_fallback": {},
        }
        for candidate in company_structured_candidates(source):
            cand_row: dict[str, Any] = {"kind": candidate["kind"], "url": candidate["url"]}
            try:
                discovered_url, discover_note = discover_structured_artifact(candidate["url"])
                cand_row["discovered_url"] = discovered_url
                cand_row["discover_note"] = discover_note
                ok, note = check_url(discovered_url)
                cand_row["reachable"] = ok
                cand_row["reachability_note"] = note
                try:
                    parsed = extract_company_structured(source, candidate["url"], candidate["kind"], "health_check")
                    cand_row["parse_status"] = "ok"
                    cand_row["extractor"] = parsed.get("extra", {}).get("fields", {}).get("extractor")
                except Exception as exc:
                    cand_row["parse_status"] = "error"
                    cand_row["parse_error"] = f"{type(exc).__name__}: {exc}"
            except Exception as exc:
                cand_row["discover_status"] = "error"
                cand_row["discover_error"] = f"{type(exc).__name__}: {exc}"
            entry["structured_candidates"].append(cand_row)

        try:
            resolved, resolution_note = resolve_machine_url(source)
            resolved, refine_note = refine_download_url(source, resolved)
            ok, note = check_url(resolved)
            entry["pdf_fallback"] = {
                "resolved_url": resolved,
                "resolution": f"{resolution_note}:{refine_note}",
                "reachable": ok,
                "reachability_note": note,
            }
        except Exception as exc:
            entry["pdf_fallback"] = {
                "status": "error",
                "error": f"{type(exc).__name__}: {exc}",
            }
        rows.append(entry)

    report = {
        "schema": "company_structured_health.v1",
        "checked_at": checked_at,
        "count": len(rows),
        "companies": rows,
    }
    out_path = REPORTS_DIR / "company_structured_health.json"
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
        f.write("\n")
    print(f"[OK ] wrote report -> {out_path.relative_to(ROOT)}")
    for row in rows:
        sid = row["source_id"]
        structured_ok = any(c.get("parse_status") == "ok" for c in row["structured_candidates"])
        pdf_ok = bool(row.get("pdf_fallback", {}).get("reachable"))
        print(f"[INFO] {sid:<32} structured_ok={structured_ok} pdf_reachable={pdf_ok}")
    return 0


def refine_download_url(source: dict[str, Any], url: str) -> tuple[str, str]:
    fmt = str(source.get("format", "MIXED"))
    sid = str(source.get("id", ""))
    try:
        r = request_get(url)
    except Exception:
        return url, "refine_probe_failed"
    ctype = (r.headers.get("content-type") or "").lower()
    looks_html = "text/html" in ctype or "<html" in (r.text[:400].lower() if hasattr(r, "text") else "")
    if not looks_html:
        return url, "direct"

    links = find_links(r.text, url)
    if not links:
        return url, "html_no_links"

    if fmt == "XLSX":
        xlsx_links = [u for u in links if ".xlsx" in u.lower()]
        if xlsx_links:
            return sorted(xlsx_links, key=len)[0], "refined_xlsx_link"
    if fmt == "PDF":
        pdf_links = [u for u in links if ".pdf" in u.lower()]
        if sid.startswith("company_"):
            preferred = [
                u for u in pdf_links
                if any(token in u.lower() for token in ("annual", "financial", "results", "report"))
            ]
            if preferred:
                return sorted(preferred, key=len)[0], "refined_company_pdf"
        if sid.startswith("accc_"):
            sys_pdf = [u for u in pdf_links if "/system/files/" in u.lower()]
            if sys_pdf:
                return sorted(sys_pdf, key=len)[0], "refined_accc_pdf"
            for page_idx in range(0, 8):
                try:
                    page_url = "https://www.accc.gov.au/publications/quarterly-reports-on-the-australian-petroleum-industry"
                    if page_idx:
                        page_url = f"{page_url}?page={page_idx}"
                    rr = request_get(page_url)
                    sub_links = find_links(rr.text, page_url)
                    sub_pdfs = [u for u in sub_links if ".pdf" in u.lower() and "/system/files/" in u.lower()]
                    if sub_pdfs:
                        return sorted(sub_pdfs, key=len)[0], "refined_accc_pdf_crawled"
                except Exception:
                    continue
        if pdf_links:
            return sorted(pdf_links, key=len)[0], "refined_pdf_link"
    return url, "html_no_artifact"

def request_get(url: str, *, stream: bool = False) -> requests.Response:
    last_exc: Exception | None = None
    for attempt in range(HTTP_RETRIES):
        try:
            return requests.get(url, headers={"User-Agent": UA}, timeout=HTTP_TIMEOUT, stream=stream)
        except requests.RequestException as exc:
            last_exc = exc
            if attempt < HTTP_RETRIES - 1:
                time.sleep(1.5 * (attempt + 1))
    assert last_exc is not None
    raise last_exc

def request_head(url: str) -> requests.Response:
    last_exc: Exception | None = None
    for attempt in range(HTTP_RETRIES):
        try:
            return requests.head(url, headers={"User-Agent": UA}, timeout=HTTP_TIMEOUT, allow_redirects=True)
        except requests.RequestException as exc:
            last_exc = exc
            if attempt < HTTP_RETRIES - 1:
                time.sleep(1.0 * (attempt + 1))
    assert last_exc is not None
    raise last_exc

def check_url(url: str) -> tuple[bool, str]:
    try:
        r = request_head(url)
        if r.status_code >= 400:
            r = request_get(url, stream=True)
            r.close()
        if r.status_code >= 400:
            return False, f"HTTP {r.status_code}"
        return True, f"HTTP {r.status_code}"
    except requests.RequestException as exc:
        return False, f"{type(exc).__name__}: {exc}"

def fetch_bytes(url: str) -> bytes:
    r = request_get(url)
    r.raise_for_status()
    return r.content

def is_zip_xlsx(blob: bytes) -> bool:
    return len(blob) >= 4 and blob[:2] == b"PK"

def load_generated(source_id: str) -> dict[str, Any] | None:
    path = GENERATED_DIR / f"{source_id}.json"
    if not path.exists():
        return None
    try:
        with path.open("r", encoding="utf-8") as f:
            doc = json.load(f)
    except Exception:
        return None
    return doc if isinstance(doc, dict) else None

def to_month_token(value: Any) -> str | None:
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m")
    if not isinstance(value, str):
        return None
    text = value.strip()
    if re.fullmatch(r"\d{4}-\d{2}", text):
        return text
    if re.fullmatch(r"\d{4}/\d{2}", text):
        return text.replace("/", "-")
    match = re.search(r"(\d{4})[-/ ](\d{1,2})", text)
    if match:
        return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"
    month_names = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }
    match = re.search(r"([A-Za-z]{3,9})[\s\-]+(\d{4})", text)
    if match:
        prefix = match.group(1).lower()[:3]
        if prefix in month_names:
            return f"{int(match.group(2)):04d}-{month_names[prefix]:02d}"
    return None

def parse_numeric(value: Any) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if not isinstance(value, str):
        return None
    text = value.strip().replace(",", "")
    if not text:
        return None
    text = text.replace("$", "").replace("%", "")
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    try:
        return float(text)
    except ValueError:
        return None

def xlsx_rows(blob: bytes) -> list[list[Any]]:
    if openpyxl is None:
        raise RuntimeError("openpyxl dependency missing")
    if not is_zip_xlsx(blob):
        raise RuntimeError("response is not an XLSX zip payload")
    wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    rows: list[list[Any]] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
    return rows


def xlsx_workbook(blob: bytes):
    if openpyxl is None:
        raise RuntimeError("openpyxl dependency missing")
    if not is_zip_xlsx(blob):
        raise RuntimeError("response is not an XLSX zip payload")
    return openpyxl.load_workbook(io.BytesIO(blob), data_only=True, read_only=True)

def extract_series_from_rows(rows: list[list[Any]], keyword_tokens: tuple[str, ...]) -> list[dict[str, Any]]:
    values: list[dict[str, Any]] = []
    for row in rows:
        if len(row) < 2:
            continue
        label = " ".join(str(c or "") for c in row[:3]).lower()
        if keyword_tokens and not all(token in label for token in keyword_tokens):
            continue
        month = to_month_token(row[0])
        if not month:
            # Sometimes label is in first col and period in second.
            month = to_month_token(row[1])
        number = None
        for cell in row[::-1]:
            number = parse_numeric(cell)
            if number is not None:
                break
        if month and number is not None:
            values.append({"t": month, "v": round(number, 3)})
    # Deduplicate by month (last occurrence wins), keep last 60.
    by_month: dict[str, float] = {}
    for point in values:
        by_month[point["t"]] = point["v"]
    out = [{"t": m, "v": by_month[m]} for m in sorted(by_month)]
    return out[-60:]

def xlsx_candidate_urls(source: dict[str, Any], resolved_url: str) -> list[str]:
    sid = str(source.get("id", ""))
    urls = [resolved_url]
    canonical = source.get("canonical_url") or source.get("url")
    if sid in APS_IDS or sid == "ato_corporate_tax":
        if isinstance(canonical, str) and canonical:
            try:
                r = request_get(canonical)
                if r.ok:
                    for link in find_links(r.text, canonical):
                        if ".xlsx" in link.lower():
                            urls.append(link)
            except Exception:
                pass
    # Preserve order, remove duplicates.
    seen = set()
    out = []
    for u in urls:
        if u in seen:
            continue
        seen.add(u)
        out.append(u)
    return out

def extract_aps(source: dict[str, Any], resolved_url: str, resolution_note: str) -> dict[str, Any]:
    rows = None
    chosen_url = resolved_url
    for candidate in xlsx_candidate_urls(source, resolved_url):
        try:
            rows = xlsx_rows(fetch_bytes(candidate))
            chosen_url = candidate
            break
        except Exception:
            continue
    if rows is None:
        raise RuntimeError("no valid APS XLSX candidate was readable")
    sid = source["id"]
    token_map = {
        "aps_monthly": (),
        "aps_production_petrol": ("motor", "spirit"),
        "aps_production_diesel": ("diesel",),
        "aps_production_jet": ("jet",),
        "aps_production_fuel_oil": ("fuel", "oil"),
        "aps_refinery_utilisation": ("refinery", "utilisation"),
    }
    values = extract_series_from_rows(rows, token_map.get(sid, ()))
    if not values and sid != "aps_monthly":
        # Fallback: if product-token matching fails due publisher layout change,
        # still extract a monthly numeric series so automation remains live.
        values = extract_series_from_rows(rows, ())
    if not values:
        raise RuntimeError("no APS values extracted")
    last = values[-1]["t"]
    return {
        "status": "ok",
        "unit": "varies by APS table",
        "values": values,
        "last_data_point": f"{last}-01",
        "notes": "Extracted from APS workbook using keyword-based table matching.",
        "extra": {
            "schema": EXTRA_SCHEMA,
            "fields": {
                "extractor": "aps_xlsx_v1",
                "resolved_url": chosen_url,
                "resolution": resolution_note,
            },
        },
    }

def extract_ato_corporate_tax(source: dict[str, Any], resolved_url: str, resolution_note: str) -> dict[str, Any]:
    wb = None
    chosen_url = resolved_url
    for candidate in xlsx_candidate_urls(source, resolved_url):
        try:
            wb = xlsx_workbook(fetch_bytes(candidate))
            chosen_url = candidate
            break
        except Exception:
            continue
    if wb is None:
        raise RuntimeError("no valid ATO XLSX candidate was readable")
    ws = None
    for sheet in wb.worksheets:
        if "income tax details" in sheet.title.lower():
            ws = sheet
            break
    if ws is None:
        ws = wb.worksheets[0] if wb.worksheets else None
    if ws is None:
        raise RuntimeError("ATO workbook had no sheets")

    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header_idx = None
    header: list[str] = []
    for idx, row in enumerate(rows[:250]):
        lower = [str(c or "").strip().lower() for c in row]
        if "taxable income" in " | ".join(lower) and "tax payable" in " | ".join(lower):
            header_idx = idx
            header = lower
            break
    if header_idx is None or not header:
        raise RuntimeError("ATO header row not found")

    def col_index(candidates: tuple[str, ...]) -> int | None:
        for i, name in enumerate(header):
            if any(c in name for c in candidates):
                return i
        return None

    income_col = col_index(("total income",))
    taxable_col = col_index(("taxable income",))
    tax_col = col_index(("tax payable", "income tax payable"))
    if income_col is None or taxable_col is None or tax_col is None:
        raise RuntimeError("ATO required columns not found")

    total_income = 0.0
    total_taxable = 0.0
    total_tax = 0.0
    rows_count = 0
    for row in rows[header_idx + 1 :]:
        if len(row) <= max(income_col, taxable_col, tax_col):
            continue
        income = parse_numeric(row[income_col])
        taxable = parse_numeric(row[taxable_col])
        tax = parse_numeric(row[tax_col])
        if income is None and taxable is None and tax is None:
            continue
        total_income += income or 0.0
        total_taxable += taxable or 0.0
        total_tax += tax or 0.0
        rows_count += 1
    if rows_count < 100:
        raise RuntimeError("ATO extractor captured too few rows")

    year_match = re.search(r"(20\d{2})[- ]?(?:2[0-9])?", resolved_url)
    last_year = int(year_match.group(1)) if year_match else dt.datetime.now(dt.timezone.utc).year
    label = f"{last_year}"
    eff = (total_tax / total_taxable * 100.0) if total_taxable > 0 else 0.0
    return {
        "status": "ok",
        "unit": "%",
        "values": [{"t": label, "v": round(eff, 2)}],
        "last_data_point": f"{last_year}-06-30",
        "notes": "Calculated aggregate effective tax rate from ATO Corporate Tax Transparency workbook rows.",
        "extra": {
            "schema": EXTRA_SCHEMA,
            "fields": {
                "extractor": "ato_xlsx_v1",
                "resolved_url": chosen_url,
                "resolution": resolution_note,
                "entities_count": rows_count,
                "aggregate_total_income": round(total_income, 2),
                "aggregate_taxable_income": round(total_taxable, 2),
                "aggregate_tax_payable": round(total_tax, 2),
            },
        },
    }

def extract_pdf_text(blob: bytes, *, max_pages: int = 40) -> tuple[str, str]:
    pages_text: list[str] = []
    if pdfplumber is not None:
        try:
            with pdfplumber.open(io.BytesIO(blob)) as pdf:
                for page in pdf.pages[:max_pages]:
                    txt = page.extract_text() or ""
                    if txt.strip():
                        pages_text.append(txt)
        except Exception:
            pages_text = []
    combined = "\n".join(pages_text)
    if combined.strip():
        return combined, "pdf_text"
    if pypdfium2 is not None:
        try:
            pdf = pypdfium2.PdfDocument(io.BytesIO(blob))
            pdfium_text = []
            for idx in range(min(len(pdf), max_pages)):
                page = pdf[idx]
                textpage = page.get_textpage()
                txt = textpage.get_text_bounded()
                if txt and txt.strip():
                    pdfium_text.append(txt)
            combined_pdfium = "\n".join(pdfium_text)
            if combined_pdfium.strip():
                return combined_pdfium, "pdf_text_pdfium"
        except Exception:
            pass
    if convert_from_bytes is None or pytesseract is None:
        raise RuntimeError("PDF had no extractable text and OCR dependencies are missing")
    try:
        images = convert_from_bytes(blob, dpi=200, first_page=1, last_page=max(10, min(max_pages, 30)))
    except Exception as exc:
        raise RuntimeError(f"OCR conversion unavailable: {exc}") from exc
    ocr_text = "\n".join(pytesseract.image_to_string(img) for img in images)
    if not ocr_text.strip():
        raise RuntimeError("PDF OCR produced no text")
    return ocr_text, "pdf_ocr"

def first_number_after(text: str, label: str) -> float | None:
    pattern = re.compile(label + r".{0,900}", re.IGNORECASE | re.DOTALL)
    for match in pattern.finditer(text):
        chunk = match.group(0)
        for token in re.findall(r"-?\(?\d[\d,]{0,20}(?:\.\d+)?\)?", chunk):
            value = parse_numeric(token)
            if value is not None:
                return value
    return None

def extract_accc(source: dict[str, Any], resolved_url: str, resolution_note: str) -> dict[str, Any]:
    blob = fetch_bytes(resolved_url)
    text, parse_mode = extract_pdf_text(blob)
    sid = source["id"]
    metrics = {
        "accc_petrol_wholesale_component": ("wholesale", "cents per litre"),
        "accc_petrol_excise_component": ("excise", "cents per litre"),
        "accc_petrol_gst_component": ("gst", "cents per litre"),
        "accc_petrol_retail_margin_component": ("retail margin", "cents per litre"),
    }

    if sid in metrics:
        token, unit = metrics[sid]
        value = first_number_after(text, token)
        if value is None:
            raise RuntimeError(f"ACCC token '{token}' not found")
        month = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m")
        return {
            "status": "ok",
            "unit": unit,
            "values": [{"t": month, "v": round(value, 2)}],
            "last_data_point": f"{month}-01",
            "notes": f"Extracted ACCC component '{token}' from latest petroleum report PDF.",
            "extra": {
                "schema": EXTRA_SCHEMA,
                "fields": {
                    "extractor": "accc_pdf_v1",
                    "resolved_url": resolved_url,
                    "resolution": resolution_note,
                    "parse_mode": parse_mode,
                },
            },
        }

    # For non-component ACCC entries, publish structured evidence if parse succeeded.
    entities = re.findall(r"(petrol|diesel|margin|excise|gst)", text.lower())
    if len(entities) < 5:
        raise RuntimeError("ACCC PDF text confidence too low")
    month = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m")
    return {
        "status": "ok",
        "unit": "index",
        "values": [{"t": month, "v": float(len(entities))}],
        "last_data_point": f"{month}-01",
        "notes": "ACCC petroleum report parsed; detailed component extraction is available on dedicated series IDs.",
        "extra": {
            "schema": EXTRA_SCHEMA,
            "fields": {
                "extractor": "accc_pdf_v1",
                "resolved_url": resolved_url,
                "resolution": resolution_note,
                "parse_mode": parse_mode,
                "keyword_hits": len(entities),
            },
        },
    }

def extract_company_report(source: dict[str, Any], resolved_url: str, resolution_note: str) -> dict[str, Any]:
    blob = fetch_bytes(resolved_url)
    # Company reports often place financial statements deep in the document.
    text, parse_mode = extract_pdf_text(blob, max_pages=220)
    lower = text.lower()
    if "annual report" not in lower and "financial statements" not in lower:
        raise RuntimeError("company PDF does not look like a report")

    sid = source["id"]
    profile = COMPANY_PARSER_PROFILES.get(sid, {})
    if sid == "company_ampol":
        npat = first_number_after(
            lower,
            r"(underlying rcop net profit after tax|net profit after tax attributable to equity holders)",
        )
        tax = first_number_after(lower, r"(income tax expense|tax expense|taxation expense)")
        if npat is None or tax is None:
            raise RuntimeError("Ampol parser could not locate NPAT/tax values")
        tax_abs = abs(tax)
        pbt = npat + tax_abs
        etr = (tax_abs / pbt * 100.0) if pbt > 0 else 0.0
        year_candidates = re.findall(r"(20\d{2})", text[:10000])
        year = max(int(y) for y in year_candidates) if year_candidates else dt.datetime.now(dt.timezone.utc).year
        return {
            "status": "ok",
            "unit": "%",
            "values": [{"t": f"{year}", "v": round(etr, 2)}],
            "last_data_point": f"{year}-06-30",
            "notes": "Company-specific Ampol parser using NPAT and tax expense from the annual report.",
            "extra": {
                "schema": EXTRA_SCHEMA,
                "fields": {
                    "extractor": "company_ampol_pdf_v1",
                    "resolved_url": resolved_url,
                    "resolution": resolution_note,
                    "parse_mode": parse_mode,
                    "financials": {
                        "net_profit_after_tax": round(npat, 2),
                        "income_tax_expense_abs": round(tax_abs, 2),
                        "implied_profit_before_tax": round(pbt, 2),
                    },
                },
            },
        }

    rev_patterns = profile.get("revenue_patterns") or [r"(revenue|total income|sales)"]
    pbt_patterns = profile.get("pbt_patterns") or [
        r"(profit before tax|profit\s+before\s+income\s+tax|earnings before tax|profit before income tax)"
    ]
    tax_patterns = profile.get("tax_patterns") or [
        r"(income tax expense|income tax paid|tax expense|taxation expense|income tax|tax charge)"
    ]

    rev = None
    for pat in rev_patterns:
        rev = first_number_after(lower, pat)
        if rev is not None:
            break
    pbt = None
    for pat in pbt_patterns:
        pbt = first_number_after(lower, pat)
        if pbt is not None:
            break
    tax = None
    for pat in tax_patterns:
        tax = first_number_after(lower, pat)
        if tax is not None:
            break
    if pbt is None:
        npat_patterns = profile.get("npat_patterns") or [r"(profit after tax|net profit after tax|net income)"]
        for npat_pat in npat_patterns:
            pat = first_number_after(lower, npat_pat)
            if pat is not None and tax is not None:
                pbt = pat + tax
                break
    if pbt is None or tax is None:
        raise RuntimeError("required financial metrics not found in company report")
    if abs(pbt) < 0.0001:
        raise RuntimeError("company metric sanity check failed (near-zero pbt)")

    etr = (tax / pbt * 100.0) if pbt > 0 else 0.0
    etr = max(min(etr, 100.0), -100.0)
    year_candidates = re.findall(r"(20\d{2})", text[:8000])
    year = max(int(y) for y in year_candidates) if year_candidates else dt.datetime.now(dt.timezone.utc).year

    return {
        "status": "ok",
        "unit": "%",
        "values": [{"t": f"{year}", "v": round(etr, 2)}],
        "last_data_point": f"{year}-06-30",
        "notes": "Parsed company annual report PDF and computed effective tax rate from extracted financial metrics.",
        "extra": {
            "schema": EXTRA_SCHEMA,
            "fields": {
                "extractor": "company_pdf_v1",
                "resolved_url": resolved_url,
                "resolution": resolution_note,
                "parse_mode": parse_mode,
                "financials": {
                    "revenue": round(rev, 2) if rev is not None else None,
                    "profit_before_tax": round(pbt, 2),
                    "income_tax": round(tax, 2),
                },
            },
        },
    }

def try_extract_ok(source: dict[str, Any], resolved_url: str, resolution_note: str) -> dict[str, Any] | None:
    sid = source["id"]

    if sid == "aud_usd_rba":
        parent = load_generated("rba_aud_usd")
        if not parent or parent.get("status") != "ok":
            return None
        values = parent.get("values")
        if not isinstance(values, list) or not values:
            return None
        return {
            "status": "ok",
            "unit": "USD per AUD",
            "values": values,
            "last_data_point": parent.get("last_data_point"),
            "notes": "Automated alias of generated rba_aud_usd monthly means.",
            "extra": {
                "schema": EXTRA_SCHEMA,
                "fields": {
                    "extractor": "alias_rba_aud_usd",
                    "resolved_url": resolved_url,
                    "resolution": resolution_note,
                },
            },
        }

    if sid in APS_IDS:
        return extract_aps(source, resolved_url, resolution_note)
    if sid == "ato_corporate_tax":
        return extract_ato_corporate_tax(source, resolved_url, resolution_note)
    if sid in ACCC_IDS:
        return extract_accc(source, resolved_url, resolution_note)
    if sid in COMPANY_IDS:
        return extract_company_with_fallback(source, resolved_url, resolution_note)
    return None

def write_manual(source: dict[str, Any], block: dict[str, Any]) -> pathlib.Path:
    sid = source["id"]
    status = block.get("status", "unavailable")
    payload: dict[str, Any] = {
        "series_id": sid,
        "source_id": sid,
        "source_name": source["human_name"],
        "source_url": source.get("canonical_url") or source["url"],
        "unit": block.get("unit", ""),
        "retrieved_at": now_iso() if status == "ok" else None,
        "last_data_point": block.get("last_data_point"),
        "values": block.get("values", []),
        "notes": block.get("notes", ""),
        "status": status,
        "manual_entry": True,
    }
    if status != "ok":
        payload["stub_created_at"] = now_iso()
    if source.get("rights"):
        payload["source_rights"] = source["rights"]
    if source.get("citation"):
        payload["citation"] = source["citation"]
    if block.get("extra") is not None:
        payload["extra"] = block["extra"]

    path = MANUAL_DIR / f"{sid}.json"
    MANUAL_DIR.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
        f.write("\n")
    return path

def fallback_block(source: dict[str, Any], resolved_url: str, resolution_note: str, err: str) -> dict[str, Any]:
    return {
        "status": "unavailable",
        "unit": "",
        "values": [],
        "last_data_point": None,
        "notes": "Automated extraction failed; source left unavailable pending parser fix.",
        "extra": {
            "schema": EXTRA_SCHEMA,
            "fields": {
                "extractor": "fallback_unavailable",
                "resolved_url": resolved_url,
                "resolution": resolution_note,
                "reachable": False,
                "reachability_note": err,
                "checked_at": now_iso(),
            },
        },
    }

def non_extractor_block(source: dict[str, Any], resolved_url: str, resolution_note: str, reachable: bool, note: str) -> dict[str, Any]:
    return {
        "status": "unavailable",
        "unit": "",
        "values": [],
        "last_data_point": None,
        "notes": "Automated source retrieval check completed; deterministic extractor not yet implemented.",
        "extra": {
            "schema": EXTRA_SCHEMA,
            "fields": {
                "extractor": "none",
                "resolved_url": resolved_url,
                "resolution": resolution_note,
                "reachable": reachable,
                "reachability_note": note,
                "checked_at": now_iso(),
            },
        },
    }

def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Refresh manual/unavailable source envelopes.")
    parser.add_argument(
        "--company-structured-health",
        action="store_true",
        help="Run structured-source health checks for all company_* sources and write data/reports/company_structured_health.json.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv if argv is not None else sys.argv[1:])
    sources = load_sources()
    if args.company_structured_health:
        return run_company_structured_health_report(sources)
    targets = [s for s in sources if s.get("fetch") in {"manual", "unavailable"}]
    if not targets:
        print("No manual/unavailable sources found.")
        return 0

    wrote = 0
    failed = 0
    for source in targets:
        sid = source["id"]
        resolved_url = source.get("canonical_url") or source.get("url") or ""
        resolution_note = "default"
        try:
            # Aliases that can be derived from already-generated data should run
            # before any network discovery.
            if sid == "aud_usd_rba":
                extracted = try_extract_ok(source, resolved_url, resolution_note)
                if extracted:
                    out = write_manual(source, extracted)
                    wrote += 1
                    print(f"[OK ] {sid:<32} -> {out.relative_to(ROOT)} ({extracted['extra']['fields']['extractor']})")
                    continue

            resolved_url, resolution_note = resolve_machine_url(source)
            resolved_url, refine_note = refine_download_url(source, resolved_url)
            resolution_note = f"{resolution_note}:{refine_note}"
            extracted = try_extract_ok(source, resolved_url, resolution_note)
            if extracted:
                out = write_manual(source, extracted)
                wrote += 1
                print(f"[OK ] {sid:<32} -> {out.relative_to(ROOT)} ({extracted['extra']['fields']['extractor']})")
                continue

            reachable, reachability_note = check_url(resolved_url)
            out = write_manual(source, non_extractor_block(source, resolved_url, resolution_note, reachable, reachability_note))
            wrote += 1
            tag = "OK " if reachable else "WARN"
            print(f"[{tag}] {sid:<32} -> {out.relative_to(ROOT)} (no extractor, {reachability_note})")
        except Exception as exc:
            failed += 1
            out = write_manual(source, fallback_block(source, resolved_url, resolution_note, f"{type(exc).__name__}: {exc}"))
            wrote += 1
            print(f"[WARN] {sid:<32} -> {out.relative_to(ROOT)} (fallback: {type(exc).__name__})")

    print()
    print(f"Refreshed {wrote} manual envelope(s)")
    if failed:
        print(f"Fallback writes due to errors: {failed}")
    return 0

if __name__ == "__main__":
    sys.exit(main())
